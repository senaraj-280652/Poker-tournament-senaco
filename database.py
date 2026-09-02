# -*- coding: utf-8 -*-
"""
Couche d'accès aux données pour le gestionnaire de tournoi de poker.
Toutes les données d'un tournoi sont stockées dans un seul fichier SQLite.
"""
import re
import sqlite3
import time
import math
import os
import glob
import shutil
import random

# =====================================================================
# Export PDF : petit utilitaire partagé par tous les export_*_pdf
# ci-dessous (Joueurs, Primes, Résultats, Gains, Synthèse par période).
# Utilise fpdf2 (police cœur "Helvetica", jeu de caractères latin-1) :
# _pdf_text() remplace les quelques symboles hors de ce jeu (tiret
# cadratin, euro) qui apparaissent dans nos en-têtes de colonnes, pour
# éviter une erreur de rendu — les exports CSV/Excel restent en Unicode
# complet, seul le PDF est concerné par cette petite simplification.
# =====================================================================

def _pdf_text(value):
    if value is None:
        return ""
    return str(value).replace("—", "-").replace("€", "EUR")


# =====================================================================
# Format français des dates (JJ/MM/AAAA) : toutes les dates/heures sont
# stockées en interne au format ISO (AAAA-MM-JJ [HH:MM:SS]), pratique
# pour le tri et les comparaisons (voir get_tournament_date,
# _read_tournament_date_ro...). Ces deux fonctions ne servent qu'à
# l'AFFICHAGE (colonnes des onglets, exports) : ne jamais les utiliser
# pour stocker ou comparer des dates.
# =====================================================================

def format_date_fr(iso_date):
    """Convertit une date "AAAA-MM-JJ" en "JJ/MM/AAAA". Renvoie la valeur
    telle quelle si elle est vide ou ne correspond pas au format attendu."""
    if not iso_date:
        return iso_date
    try:
        y, m, d = iso_date.split("-")
        if len(y) == 4 and len(m) == 2 and len(d) == 2:
            return f"{d}/{m}/{y}"
    except (ValueError, AttributeError):
        pass
    return iso_date


def format_datetime_fr(iso_dt):
    """Convertit "AAAA-MM-JJ HH:MM:SS" en "JJ/MM/AAAA HH:MM:SS". Renvoie la
    valeur telle quelle si elle est vide ou ne correspond pas au format
    attendu."""
    if not iso_dt:
        return iso_dt
    try:
        date_part, time_part = iso_dt.split(" ", 1)
        formatted_date = format_date_fr(date_part)
        if formatted_date == date_part:
            return iso_dt
        return f"{formatted_date} {time_part}"
    except (ValueError, AttributeError):
        return iso_dt


def _pdf_fit_font_size(pdf, texts, col_width, bold=False, max_size=9, min_size=5):
    """Plus grande taille de police (entre `min_size` et `max_size`) à
    laquelle chacun de `texts` tient dans une colonne de largeur
    `col_width` (avec 2mm de marge) — évite que des en-têtes/valeurs longs
    ne débordent sur la colonne suivante quand il y a beaucoup de
    colonnes. Laisse la police active sur ce choix en sortie."""
    style = "B" if bold else ""
    for size in range(max_size, min_size - 1, -1):
        pdf.set_font("Helvetica", style, size)
        if all(pdf.get_string_width(t) <= col_width - 2 for t in texts):
            return size
    pdf.set_font("Helvetica", style, min_size)
    return min_size


def _write_pdf_table(path, title, subtitle_lines, headers, rows):
    """Génère un PDF simple (titre, sous-titres, puis un tableau) à partir
    de lignes déjà calculées (mêmes valeurs que pour les exports CSV/
    Excel). `subtitle_lines` : liste de lignes de texte optionnelles sous
    le titre (ex : entrées/prize pool). Paysage automatique au-delà de 6
    colonnes, pour laisser assez de place à chacune. La taille de police
    des en-têtes et des valeurs s'ajuste automatiquement (voir
    _pdf_fit_font_size) pour ne jamais déborder d'une colonne."""
    from fpdf import FPDF

    orientation = "L" if len(headers) > 6 else "P"
    pdf = FPDF(orientation=orientation, unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _pdf_text(title), border=0)
    pdf.ln(10)
    if subtitle_lines:
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(90, 90, 90)
        for line in subtitle_lines:
            pdf.cell(0, 6, _pdf_text(line), border=0)
            pdf.ln(6)
        pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    avail_width = pdf.w - pdf.l_margin - pdf.r_margin
    col_width = avail_width / max(1, len(headers))
    row_height = 7

    header_texts = [_pdf_text(h) for h in headers]
    _pdf_fit_font_size(pdf, header_texts, col_width, bold=True)
    pdf.set_fill_color(31, 78, 36)
    pdf.set_text_color(255, 255, 255)
    for h in header_texts:
        pdf.cell(col_width, row_height + 1, h, border=1, align="C", fill=True)
    pdf.ln(row_height + 1)

    body_texts = [_pdf_text(c) for row in rows for c in row] or [""]
    _pdf_fit_font_size(pdf, body_texts, col_width, bold=False)
    pdf.set_text_color(0, 0, 0)
    fill_toggle = False
    for row in rows:
        if fill_toggle:
            pdf.set_fill_color(247, 241, 227)
        else:
            pdf.set_fill_color(255, 255, 255)
        for cell in row:
            pdf.cell(col_width, row_height, _pdf_text(cell), border=1, align="C", fill=True)
        pdf.ln(row_height)
        fill_toggle = not fill_toggle

    pdf.output(path)
    return path


SCHEMA = """
CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS tables_pk (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    max_seats INTEGER NOT NULL DEFAULT 9,
    is_active INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS players (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    buyin_count INTEGER NOT NULL DEFAULT 1,
    rebuy_count INTEGER NOT NULL DEFAULT 0,
    addon_count INTEGER NOT NULL DEFAULT 0,
    chips INTEGER NOT NULL DEFAULT 0,
    table_id INTEGER,
    seat INTEGER,
    status TEXT NOT NULL DEFAULT 'active',   -- active | eliminated | withdrawn
    place INTEGER,
    elim_time TEXT,
    bounty INTEGER NOT NULL DEFAULT 0,       -- prime actuellement portée par ce joueur
    bounty_won INTEGER NOT NULL DEFAULT 0,   -- cumul des primes empochées (en cash)
    kills INTEGER NOT NULL DEFAULT 0,        -- nb de joueurs éliminés par ce joueur (prime de bounty en points)
    club TEXT NOT NULL DEFAULT '',           -- club du joueur POUR CE TOURNOI (copié du répertoire à l'ajout, voir roster.py)
    FOREIGN KEY(table_id) REFERENCES tables_pk(id)
);

CREATE TABLE IF NOT EXISTS blind_levels (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    level_order INTEGER NOT NULL,
    small_blind INTEGER NOT NULL,
    big_blind INTEGER NOT NULL,
    ante INTEGER NOT NULL DEFAULT 0,
    duration_minutes INTEGER NOT NULL,
    is_break INTEGER NOT NULL DEFAULT 0,
    break_label TEXT
);

CREATE TABLE IF NOT EXISTS payout_structure (
    place INTEGER PRIMARY KEY,
    percentage REAL NOT NULL
);

CREATE TABLE IF NOT EXISTS seat_moves (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    player_name TEXT NOT NULL,
    old_table_name TEXT,
    old_seat INTEGER,
    new_table_name TEXT,
    new_seat INTEGER,
    moved_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS bounty_events (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eliminated_name TEXT NOT NULL,
    eliminator_name TEXT,
    amount_won INTEGER NOT NULL,
    added_to_eliminator_bounty INTEGER NOT NULL DEFAULT 0,
    event_time TEXT NOT NULL
);
"""

DEFAULT_SETTINGS = {
    "tournament_name": "Nouveau tournoi",
    "buyin_amount": "50",
    "rebuy_amount": "50",
    "addon_amount": "50",
    "starting_chips": "10000",
    "rebuy_chips": "10000",
    "addon_chips": "10000",
    "max_seats_per_table": "9",
    "min_players_per_table": "4",
    "break_duration_minutes": "15",
    "movement_signal_duration_ms": "300",
    "highlight_duration_minutes": "5",
    "rake_percent": "0",
    "bounty_amount": "0",
    "pko_mode": "0",
    "pko_cash_percent": "50",
    "current_level_order": "1",
    "level_start_epoch": "0",
    "is_paused": "1",
    "paused_accum_seconds": "0",
    "clock_started": "0",
    "tournament_date": "",  # AAAA-MM-JJ, fixée à la création (voir get_tournament_date)
    "tournament_start_epoch": "0",  # fixé au tout premier "Démarrer" (voir App._clock_resume)
    "tournament_end_epoch": "0",  # fixé quand il ne reste plus qu'1 joueur actif (voir eliminate_player)
}


def ranking_points(place, n_players, flat_value=0):
    """Valeur en points de la prime de classement pour un rang `place`
    parmi `n_players` joueurs au total dans le tournoi. Si `flat_value`
    (réglage manuel, non nul) est fourni, il est utilisé tel quel pour
    tout le monde ; sinon on applique la formule 100×√N/P, pensée pour ne
    pas distribuer une masse de points disproportionnée sur les petites
    tables (elle grandit avec le nombre de joueurs et décroît avec un
    rang moins bon)."""
    if flat_value:
        return flat_value
    if not place or place <= 0 or n_players <= 0:
        return 0
    return round(100 * math.sqrt(n_players) / place)


def bounty_unit_value(n_players, flat_value=0):
    """Valeur en points d'un bounty (un joueur éliminé) dans un tournoi de
    `n_players` joueurs au total. Si `flat_value` (réglage manuel, non nul)
    est fourni, il est utilisé tel quel pour tout le monde ; sinon on
    applique la formule 10×√N — plus le champ est grand, plus éliminer un
    adversaire y est statistiquement difficile, donc plus le bounty
    rapporte."""
    if flat_value:
        return flat_value
    if n_players <= 0:
        return 0
    return round(10 * math.sqrt(n_players))


# Convention "table finale" (voir rebalance_tables) : une fois qu'il ne
# reste plus que ce nombre de joueurs actifs ou moins, ils sont toujours
# regroupés sur UNE SEULE table, quitte à dépasser ponctuellement le
# réglage "Nombre de sièges par table" s'il est plus petit (ex. 8) — un
# vrai tournoi ne scinde jamais les tout derniers joueurs entre deux
# tables alors qu'ils tiendraient sur une seule table finale.
FINAL_TABLE_MAX_SEATS = 10


class Database:
    def __init__(self, path, read_only=False):
        """`read_only=True` : pour une simple consultation (Lobby SNG,
        synthèse par période...) d'un fichier .tournoi potentiellement
        déjà ouvert par une autre fenêtre/processus en ce moment même
        (voir _read_tournament_date_ro plus bas dans ce module, même
        principe). Ouvre en mode URI "ro" et saute la création/migration
        du schéma (executescript + _migrate + _init_defaults + commit,
        qui prennent chacun un verrou d'écriture) : un fichier .tournoi
        existant a forcément déjà tout ça en place, inutile de le refaire
        juste pour lire. Sous Windows en particulier, répéter ces
        écritures à chaque rafraîchissement du Lobby (toutes les 4s, sur
        chaque fichier du dossier) entrait en conflit avec les écritures
        de la fenêtre qui a ce même tournoi ouvert, et le fichier
        disparaissait alors silencieusement de la liste (exception
        avalée par l'appelant) le temps du conflit."""
        self.path = path
        if read_only:
            self.conn = sqlite3.connect(f"file:{os.path.abspath(path)}?mode=ro", uri=True)
            self.conn.row_factory = sqlite3.Row
            return
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.conn.executescript(SCHEMA)
        self._migrate()
        self._init_defaults()
        self.conn.commit()

    def _migrate(self):
        """Ajoute les colonnes apparues après la création initiale du
        fichier .tournoi (les anciens fichiers n'ont pas 'bounty' /
        'bounty_won' / 'kills' sur la table players)."""
        cols = {row["name"] for row in self.conn.execute("PRAGMA table_info(players)")}
        if "bounty" not in cols:
            self.conn.execute("ALTER TABLE players ADD COLUMN bounty INTEGER NOT NULL DEFAULT 0")
        if "bounty_won" not in cols:
            self.conn.execute("ALTER TABLE players ADD COLUMN bounty_won INTEGER NOT NULL DEFAULT 0")
        if "kills" not in cols:
            self.conn.execute("ALTER TABLE players ADD COLUMN kills INTEGER NOT NULL DEFAULT 0")
        if "elim_round" not in cols:
            self.conn.execute("ALTER TABLE players ADD COLUMN elim_round INTEGER")
        if "eliminated_by_name" not in cols:
            self.conn.execute("ALTER TABLE players ADD COLUMN eliminated_by_name TEXT")
        if "club" not in cols:
            self.conn.execute("ALTER TABLE players ADD COLUMN club TEXT NOT NULL DEFAULT ''")

    # ---------- init ----------
    def _init_defaults(self):
        cur = self.conn.cursor()
        for k, v in DEFAULT_SETTINGS.items():
            cur.execute(
                "INSERT OR IGNORE INTO settings(key, value) VALUES (?, ?)", (k, v)
            )
        cur.execute("SELECT COUNT(*) c FROM blind_levels")
        if cur.fetchone()["c"] == 0:
            from structures import default_blind_structure
            self.set_blind_structure(default_blind_structure())
        cur.execute("SELECT COUNT(*) c FROM payout_structure")
        if cur.fetchone()["c"] == 0:
            self.set_payout_structure({1: 100.0})
        cur.execute("SELECT COUNT(*) c FROM tables_pk")
        if cur.fetchone()["c"] == 0:
            self.add_table("Table 1")

    # ---------- settings ----------
    def get_setting(self, key, default=None):
        row = self.conn.execute(
            "SELECT value FROM settings WHERE key=?", (key,)
        ).fetchone()
        return row["value"] if row else default

    def get_setting_int(self, key, default=0):
        v = self.get_setting(key)
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return default

    def get_setting_float(self, key, default=0.0):
        v = self.get_setting(key)
        try:
            return float(v)
        except (TypeError, ValueError):
            return default

    def set_setting(self, key, value):
        self.conn.execute(
            "INSERT INTO settings(key, value) VALUES (?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, str(value)),
        )
        self.conn.commit()

    def get_tournament_date(self):
        """Date du tournoi (AAAA-MM-JJ), utilisée pour les synthèses par
        période. C'est la date fixée à la création du tournoi si elle est
        connue ; sinon (fichiers créés avant l'existence de ce paramètre),
        on retombe sur la date de création du fichier .tournoi lui-même
        (st_birthtime si disponible - macOS/BSD -, sinon la date de
        dernière modification). On évite volontairement de se baser sur
        la date de modification quand la vraie date de création est
        connue : le simple fait d'ouvrir un ancien fichier (migrations de
        réglages) le "touche" et modifierait sinon sa date à chaque
        utilisation."""
        d = self.get_setting("tournament_date", "")
        if d:
            return d
        try:
            st = os.stat(self.path)
            ts = getattr(st, "st_birthtime", None)
            if ts is None:
                ts = st.st_mtime
            return time.strftime("%Y-%m-%d", time.localtime(ts))
        except OSError:
            return ""

    def set_settings(self, mapping):
        for k, v in mapping.items():
            self.conn.execute(
                "INSERT INTO settings(key, value) VALUES (?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                (k, str(v)),
            )
        self.conn.commit()

    # ---------- tables ----------
    def set_all_tables_max_seats(self, new_max):
        """Applique un nouveau nombre de sièges par table à TOUTES les
        tables existantes (actives ou fermées), pour que le changement du
        paramètre prenne effet immédiatement, y compris en cours de
        tournoi. À appeler suivi d'un rebalance_tables()."""
        new_max = max(2, int(new_max))
        self.conn.execute("UPDATE tables_pk SET max_seats=?", (new_max,))
        self.conn.commit()

    def _table_display_number(self, table_row):
        """Numéro affiché d'une table (ex : 9 pour "Table 9"), utilisé pour
        décider quelle table fermer en premier lors d'une consolidation
        (voir rebalance_tables) — à préférer à l'id interne, qui peut
        diverger du numéro affiché sur un vieux fichier .tournoi (voir
        commentaire dans rebalance_tables). Retombe sur l'id si le nom ne
        se termine pas par un nombre (table renommée manuellement)."""
        m = re.search(r"(\d+)\s*$", table_row["name"] or "")
        return int(m.group(1)) if m else table_row["id"]

    def _open_or_reopen_table(self):
        """Fournit une table supplémentaire quand aucune table active n'a
        de place (voir _seat_player et rebalance_tables) : réactive
        d'abord la table FERMÉE dont le numéro affiché est le plus bas
        s'il en existe une, plutôt que d'en créer systématiquement une
        toute nouvelle avec un numéro plus haut. Sans ça, une phase de
        consolidation (tables fermées en fin de tournoi) suivie d'un
        besoin de place retrouvé (joueur réintégré, ou ajouté après coup)
        rouvrait toujours une table de numéro croissant, laissant des
        trous dans la numérotation affichée (ex : Table 1, 8, 9 restantes
        au lieu de Table 1, 2, 3 — un joueur avait fini par signaler cette
        numérotation en dents de scie comme un bug). Renvoie la ligne de
        la table (réactivée ou nouvellement créée)."""
        closed = [t for t in self.list_tables(active_only=False) if t["is_active"] == 0]
        if closed:
            closed.sort(key=self._table_display_number)
            t = closed[0]
            self.conn.execute("UPDATE tables_pk SET is_active=1 WHERE id=?", (t["id"],))
            self.conn.commit()
            return self.conn.execute("SELECT * FROM tables_pk WHERE id=?", (t["id"],)).fetchone()
        new_id = self.add_table()
        return self.conn.execute("SELECT * FROM tables_pk WHERE id=?", (new_id,)).fetchone()

    def add_table(self, name=None):
        max_seats = self.get_setting_int("max_seats_per_table", 9)
        if name is None:
            # Compte TOUTES les tables jamais créées (actives ou fermées) :
            # se baser uniquement sur les tables actives ferait repartir la
            # numérotation en arrière après une fermeture, et créerait des
            # doublons de nom (ex : "Table 2" utilisé deux fois).
            n = self.conn.execute("SELECT COUNT(*) c FROM tables_pk").fetchone()["c"]
            name = f"Table {n + 1}"
        cur = self.conn.execute(
            "INSERT INTO tables_pk(name, max_seats, is_active) VALUES (?, ?, 1)",
            (name, max_seats),
        )
        self.conn.commit()
        return cur.lastrowid

    def list_tables(self, active_only=True):
        q = "SELECT * FROM tables_pk"
        if active_only:
            q += " WHERE is_active=1"
        q += " ORDER BY id"
        return self.conn.execute(q).fetchall()

    def close_table(self, table_id):
        self.conn.execute(
            "UPDATE tables_pk SET is_active=0 WHERE id=?", (table_id,)
        )
        self.conn.commit()

    # ---------- players ----------
    def list_players(self, status=None):
        q = "SELECT * FROM players"
        params = ()
        if status:
            q += " WHERE status=?"
            params = (status,)
        q += " ORDER BY table_id, seat"
        return self.conn.execute(q, params).fetchall()

    def get_player(self, player_id):
        return self.conn.execute(
            "SELECT * FROM players WHERE id=?", (player_id,)
        ).fetchone()

    def find_active_conflict(self, player_name):
        """Renvoie le chemin du fichier .tournoi (dans le même dossier que
        celui-ci, non récursif, et daté du même jour — voir plus bas) où
        `player_name` est actuellement un joueur actif, ou None s'il n'y
        en a pas. Sert à empêcher d'inscrire par erreur un même joueur sur
        deux tournois en cours à la fois (ex : deux Sit & Go simultanés
        dans la même salle). Les autres fichiers ne sont consultés qu'en
        lecture seule. Si ce tournoi n'a pas encore été sauvegardé sur
        disque, aucune vérification n'est possible (renvoie None).

        Seuls les tournois datés du même jour que celui-ci sont comparés :
        un ancien fichier abandonné (jamais terminé, un autre jour) ne
        doit pas bloquer indéfiniment un joueur qui n'est plus vraiment
        "en jeu ailleurs"."""
        if not self.path or not os.path.exists(self.path) or not player_name.strip():
            return None
        folder = os.path.dirname(os.path.abspath(self.path)) or "."
        today = self.get_tournament_date()
        name_lower = player_name.strip().lower()
        for path in find_tournament_files(folder, recursive=False):
            if os.path.abspath(path) == os.path.abspath(self.path):
                continue
            if today and _read_tournament_date_ro(path) != today:
                continue
            if _player_active_in_file(path, name_lower):
                return path
        return None

    def add_player(self, name, club=""):
        """`club` : copié dans ce tournoi au moment de l'ajout (voir
        roster.get_club côté appelant) — n'est ensuite plus synchronisé
        avec le répertoire si celui-ci change, ce tournoi garde la photo
        du club tel qu'il était à l'inscription."""
        starting_chips = self.get_setting_int("starting_chips", 10000)
        bounty_amount = self.get_setting_int("bounty_amount", 0)
        cur = self.conn.execute(
            "INSERT INTO players(name, buyin_count, rebuy_count, addon_count, "
            "chips, status, bounty, club) VALUES (?, 1, 0, 0, ?, 'active', ?, ?)",
            (name, starting_chips, bounty_amount, (club or "").strip()),
        )
        player_id = cur.lastrowid
        self.conn.commit()
        self._seat_player(player_id)
        self.rebalance_tables(record_moves=False)
        return player_id

    def rebuy_player(self, player_id):
        chips = self.get_setting_int("rebuy_chips", 10000)
        bounty_amount = self.get_setting_int("bounty_amount", 0)
        self.conn.execute(
            "UPDATE players SET rebuy_count = rebuy_count + 1, "
            "chips = chips + ?, bounty = bounty + ? WHERE id=?",
            (chips, bounty_amount, player_id),
        )
        self.conn.commit()

    def addon_player(self, player_id):
        chips = self.get_setting_int("addon_chips", 10000)
        self.conn.execute(
            "UPDATE players SET addon_count = addon_count + 1, "
            "chips = chips + ? WHERE id=?",
            (chips, player_id),
        )
        self.conn.commit()

    def rename_player(self, player_id, new_name):
        new_name = new_name.strip()
        if not new_name:
            return
        self.conn.execute(
            "UPDATE players SET name=? WHERE id=?", (new_name, player_id)
        )
        self.conn.commit()

    def set_player_club(self, player_id, club):
        """Corrige le club (POUR CE TOURNOI, voir add_player) d'un joueur
        déjà inscrit — ex. club mal renseigné/absent à l'ajout."""
        self.conn.execute(
            "UPDATE players SET club=? WHERE id=?", ((club or "").strip(), player_id)
        )
        self.conn.commit()

    def set_purchase_counts(self, player_id, buyin_count, rebuy_count, addon_count):
        """Corrige manuellement les compteurs buy-in / rebuy / add-on d'un
        joueur (utile en cas d'erreur de saisie), sans toucher aux chips."""
        self.conn.execute(
            "UPDATE players SET buyin_count=?, rebuy_count=?, addon_count=? WHERE id=?",
            (max(0, int(buyin_count)), max(0, int(rebuy_count)), max(0, int(addon_count)), player_id),
        )
        self.conn.commit()

    def set_chips(self, player_id, chips):
        self.conn.execute(
            "UPDATE players SET chips=? WHERE id=?", (max(0, int(chips)), player_id)
        )
        self.conn.commit()

    def eliminate_player(self, player_id, eliminated_by_id=None):
        """Élimine un joueur. Si `eliminated_by_id` est fourni, l'éliminateur
        voit son compteur de bounty (kills, prime de bounty en points —
        voir get_bounty_bonuses) incrémenté de 1, quel que soit l'ancien
        mécanisme de bounty en €. Si en plus le joueur éliminé portait une
        prime (bounty €), celle-ci est versée à l'éliminateur : intégralement
        en mode classique, ou selon le partage PKO (une partie en cash
        immédiat, le reste ajouté à la prime de l'éliminateur) en mode
        progressif. Enregistre aussi, pour l'onglet Joueurs, le round et le
        nom de l'éliminateur (indépendamment de tout bounty en €)."""
        active = self.list_players(status="active")
        place = len(active)  # ce joueur prend la place n° (nb d'actifs restants)
        eliminated = self.get_player(player_id)
        now = time.strftime("%Y-%m-%d %H:%M:%S")
        current_round = self.get_current_round_number()
        eliminator_row = self.get_player(eliminated_by_id) if eliminated_by_id else None
        eliminator_name = eliminator_row["name"] if eliminator_row else None

        self.conn.execute(
            "UPDATE players SET status='eliminated', place=?, elim_time=?, "
            "elim_round=?, eliminated_by_name=?, table_id=NULL, seat=NULL WHERE id=?",
            (place, now, current_round, eliminator_name, player_id),
        )

        if eliminated_by_id:
            self.conn.execute(
                "UPDATE players SET kills = kills + 1 WHERE id=?", (eliminated_by_id,)
            )

        if eliminated_by_id and eliminated and eliminated["bounty"] > 0:
            bounty = eliminated["bounty"]
            eliminator = self.get_player(eliminated_by_id)
            pko_mode = self.get_setting_int("pko_mode", 0) == 1
            if pko_mode:
                cash_pct = self.get_setting_int("pko_cash_percent", 50)
                cash_part = round(bounty * cash_pct / 100)
                grow_part = bounty - cash_part
            else:
                cash_part = bounty
                grow_part = 0
            self.conn.execute(
                "UPDATE players SET bounty_won = bounty_won + ?, bounty = bounty + ? "
                "WHERE id=?",
                (cash_part, grow_part, eliminated_by_id),
            )
            self.conn.execute("UPDATE players SET bounty=0 WHERE id=?", (player_id,))
            self.conn.execute(
                "INSERT INTO bounty_events(eliminated_name, eliminator_name, "
                "amount_won, added_to_eliminator_bounty, event_time) "
                "VALUES (?,?,?,?,?)",
                (eliminated["name"], eliminator["name"] if eliminator else "?",
                 cash_part, grow_part, now),
            )

        self.conn.commit()

        # Fige l'heure de fin dès qu'il ne reste plus qu'1 joueur actif
        # (le vainqueur) — sert à figer l'affichage "Durée" du chrono
        # projecteur au lieu de continuer à défiler après la fin de la
        # partie, et à y afficher "Partie terminée" (voir get_stats()).
        if len(self.list_players(status="active")) <= 1:
            self.set_setting("tournament_end_epoch", int(time.time()))

        return self.rebalance_tables(record_moves=True)

    def withdraw_player(self, player_id):
        """Retire un joueur de la liste active sans lui attribuer de place
        au classement (forfait / inscription annulée), contrairement à
        eliminate_player. Libère sa table/siège comme une élimination."""
        self.conn.execute(
            "UPDATE players SET status='withdrawn', place=NULL, elim_time=?, "
            "table_id=NULL, seat=NULL WHERE id=?",
            (time.strftime("%Y-%m-%d %H:%M:%S"), player_id),
        )
        # Ce joueur était compté dans "active" (voir eliminate_player :
        # place = nb d'actifs restants au moment de l'élimination) pour
        # TOUT joueur déjà éliminé jusqu'ici, tant qu'il n'avait pas
        # encore forfait. Comme un forfait ne prend jamais de place au
        # classement (place reste NULL ci-dessus), il faut le retirer
        # rétroactivement du calcul : le nombre réel de participants
        # classés diminue d'une unité, donc chaque rang déjà attribué
        # doit être décalé d'une place vers le haut (ex : 3e devient 2e)
        # pour ne pas laisser de trou dans le classement final.
        self.conn.execute(
            "UPDATE players SET place = place - 1 WHERE status='eliminated' AND place IS NOT NULL"
        )
        self.conn.commit()
        return self.rebalance_tables(record_moves=False)

    def reinstate_player(self, player_id):
        starting_chips = self.get_setting_int("starting_chips", 10000)
        bounty_amount = self.get_setting_int("bounty_amount", 0)
        was_withdrawn = self.get_player(player_id)["status"] == "withdrawn"
        self.conn.execute(
            "UPDATE players SET status='active', place=NULL, elim_time=NULL, "
            "chips=?, bounty=? WHERE id=?",
            (starting_chips, bounty_amount, player_id),
        )
        if was_withdrawn:
            # Symétrique du décalage fait dans withdraw_player : ce joueur
            # réintègre le décompte des participants classés, chaque rang
            # déjà attribué redescend donc d'une place (ex : 2e redevient
            # 3e). Sans effet s'il était éliminé (pas forfait) : son
            # départ n'avait alors jamais touché aux rangs des autres.
            self.conn.execute(
                "UPDATE players SET place = place + 1 WHERE status='eliminated' AND place IS NOT NULL"
            )
        self.conn.commit()
        self._seat_player(player_id)
        # Réintégrer un joueur peut faire repasser le nombre d'actifs
        # au-dessus de 1 : la partie n'est alors plus terminée, on efface
        # l'heure de fin figée (voir eliminate_player) pour que "Durée"
        # se remette à compter sur le chrono projecteur.
        if len(self.list_players(status="active")) > 1:
            self.set_setting("tournament_end_epoch", 0)
        return self.rebalance_tables(record_moves=False)

    def delete_player(self, player_id):
        self.conn.execute("DELETE FROM players WHERE id=?", (player_id,))
        self.conn.commit()
        return self.rebalance_tables(record_moves=False)

    # ---------- seating / balancing ----------
    def _seat_player(self, player_id):
        """Assoit un joueur à la table la moins remplie, sur le premier siège libre."""
        tables = self.list_tables()
        if not tables:
            self._open_or_reopen_table()
            tables = self.list_tables()
        best_table = None
        best_count = None
        for t in tables:
            occ = self.conn.execute(
                "SELECT COUNT(*) c FROM players WHERE table_id=? AND status='active'",
                (t["id"],),
            ).fetchone()["c"]
            if occ < t["max_seats"] and (best_count is None or occ < best_count):
                best_table, best_count = t, occ
        if best_table is None:
            best_table = self._open_or_reopen_table()
        taken = {
            r["seat"]
            for r in self.conn.execute(
                "SELECT seat FROM players WHERE table_id=? AND status='active'",
                (best_table["id"],),
            )
        }
        seat = 1
        while seat in taken:
            seat += 1
        self.conn.execute(
            "UPDATE players SET table_id=?, seat=? WHERE id=?",
            (best_table["id"], seat, player_id),
        )
        self.conn.commit()

    def rebalance_tables(self, record_moves=False):
        """Rééquilibre les tables actives : comble les sièges vides en déplaçant
        des joueurs des tables les plus pleines, et ferme les tables devenues
        inutiles quand le nombre de joueurs restants tient sur moins de
        tables — toujours en partant du numéro de table le plus haut (les
        tables gardent leur numéro toute la partie, jamais renumérotées :
        voir plus bas). Si record_moves est vrai, archive chaque déplacement réel (ancienne
        table/siège -> nouvelle table/siège) dans l'historique des
        mouvements (onglet Mouvements) — ce n'est le cas que pour les
        rééquilibrages déclenchés par une élimination de joueur. Renvoie
        dans tous les cas la liste des mouvements effectués.

        Ne fait rien quand il reste 0 ou 1 joueur actif : à 1 seul joueur
        actif, le tournoi est terminé (il n'y a plus personne à équilibrer
        entre tables) — sans ce garde-fou, le vainqueur pouvait se
        retrouver déplacé vers une table "consolidée" au tout dernier
        rééquilibrage et apparaître à tort dans l'historique des
        mouvements alors que la partie est finie."""
        active_players = [
            dict(p) for p in self.list_players(status="active")
        ]
        n_active = len(active_players)
        if n_active <= 1:
            return []

        before_state = {p["id"]: (p["table_id"], p["seat"]) for p in active_players}

        max_seats = self.get_setting_int("max_seats_per_table", 9)
        min_players = self.get_setting_int("min_players_per_table", 4)
        tables = list(self.list_tables())
        n_tables_needed = max(1, math.ceil(n_active / max_seats))

        # En dessous du seuil minimum de joueurs par table, on regroupe
        # davantage (une table de plus en moins) tant que c'est possible
        # sans dépasser le nombre de sièges disponibles par table.
        if min_players > 1:
            while (
                n_tables_needed > 1
                and n_active / n_tables_needed < min_players
                and n_active <= (n_tables_needed - 1) * max_seats
            ):
                n_tables_needed -= 1

        # Convention "table finale" (voir FINAL_TABLE_MAX_SEATS) : au
        # poker, la toute dernière table peut accueillir jusqu'à 10
        # joueurs même si "Nombre de sièges par table" est réglé plus bas
        # (ex. 8) — on ne scinde jamais les derniers joueurs entre deux
        # tables alors qu'ils tiendraient sur une seule table finale.
        if n_active <= FINAL_TABLE_MAX_SEATS:
            n_tables_needed = 1

        # Ouvre des tables supplémentaires si le nombre de sièges disponibles
        # ne suffit plus (ex : réduction du nombre de sièges par table en
        # cours de tournoi).
        if len(tables) < n_tables_needed:
            for _ in range(n_tables_needed - len(tables)):
                self._open_or_reopen_table()
            tables = list(self.list_tables())

        # Si consolider sur une seule table finale dépasse le nombre de
        # sièges normalement configuré (cas ci-dessus), relève
        # ponctuellement la capacité de CETTE table pour qu'elle puisse
        # réellement tous les accueillir — sinon _seat_player refuserait
        # de la remplir au-delà de son max_seats actuel et rouvrirait une
        # table à la place, annulant la fusion voulue.
        if n_tables_needed == 1 and n_active > max_seats:
            survivor = min(tables, key=self._table_display_number)
            if survivor["max_seats"] < n_active:
                self.conn.execute(
                    "UPDATE tables_pk SET max_seats=? WHERE id=?",
                    (n_active, survivor["id"]),
                )
                self.conn.commit()
                tables = list(self.list_tables())
        elif any(t["max_seats"] != max_seats for t in tables):
            # Repli : plusieurs tables sont de nouveau nécessaires (ex :
            # d'autres joueurs se sont inscrits depuis une fusion "table
            # finale" ci-dessus) — sans ça, la table déjà ponctuellement
            # élargie garderait pour toujours une capacité différente des
            # autres, sans raison apparente une fois la fusion plus
            # nécessaire.
            self.conn.execute("UPDATE tables_pk SET max_seats=?", (max_seats,))
            self.conn.commit()
            tables = list(self.list_tables())

        # Ferme les tables en trop — toujours en partant du numéro le PLUS
        # HAUT (comme dans un vrai tournoi : les tables sont pré-numérotées
        # à leur installation et gardent ce numéro toute la partie ; on
        # regroupe progressivement les joueurs vers les tables 1, 2, 3...
        # jusqu'à la table finale n°1, jamais l'inverse). Trier par nombre
        # de joueurs (vider la plus petite d'abord) déplacerait moins de
        # monde en moyenne, mais casserait cette convention : une table
        # basse pourrait fermer avant une table haute plus vide, ce qui ne
        # correspond à aucune pratique réelle de gestion de tournoi et
        # obligeait jusqu'ici à renuméroter les tables restantes (source
        # de confusion dans l'historique des mouvements — deux tables
        # différentes affichées sous le même nom à des moments différents).
        if len(tables) > n_tables_needed:
            occ_by_table = {}
            for p in active_players:
                occ_by_table.setdefault(p["table_id"], []).append(p)
            # Trié par NUMÉRO affiché (extrait du nom "Table N"), pas par id
            # interne : les deux coïncident normalement (add_table nomme
            # toujours la nouvelle table d'après le nombre total de tables
            # jamais créées), mais un fichier .tournoi hérité d'une
            # version antérieure à la suppression de l'ancienne
            # renumérotation pouvait avoir des id et des numéros affichés
            # décorrélés — d'où par exemple "Table 1, Table 8, Table 9"
            # restantes au lieu de "Table 1, Table 2, Table 3" quand on se
            # fie à l'id brut plutôt qu'au numéro réellement affiché.
            tables_sorted = sorted(tables, key=lambda t: -self._table_display_number(t))
            to_close = tables_sorted[: len(tables) - n_tables_needed]
            # Regroupe TOUS les joueurs évincés de TOUTES les tables fermées
            # dans CE MÊME passage (ex : fusion directe vers la table
            # finale, qui ferme souvent plusieurs tables d'un coup) en une
            # seule liste, plutôt que de les réasseoir table fermée par
            # table fermée : sans ça, les joueurs de la première table
            # fermée occuperaient systématiquement les places les plus
            # "précoces", un biais détectable même si chaque groupe était
            # mélangé séparément.
            players_to_move = []
            for t in to_close:
                players_to_move.extend(occ_by_table.get(t["id"], []))
                self.close_table(t["id"])
            for p in players_to_move:
                self.conn.execute(
                    "UPDATE players SET table_id=NULL, seat=NULL WHERE id=?",
                    (p["id"],),
                )
            self.conn.commit()
            # Cassage de table (contrairement au simple équilibrage
            # ci-dessous, laissé inchangé) : répartition ALÉATOIRE des
            # joueurs évincés sur les places disponibles des tables
            # restantes. _seat_player() choisit toujours la table la moins
            # remplie puis le premier siège libre — une séquence de places
            # entièrement déterminée par l'état d'occupation courant,
            # jamais par l'identité du joueur passé en argument. Mélanger
            # l'ORDRE des joueurs avant de les réasseoir un par un dans
            # cette même séquence de places (déjà équilibrée) équivaut donc
            # à une bijection aléatoire uniforme joueur -> place, sans
            # toucher à _seat_player() elle-même (qui doit continuer à
            # garantir des effectifs équilibrés).
            random.shuffle(players_to_move)
            for p in players_to_move:
                self._seat_player(p["id"])

        # Ré-équilibre : déplace un joueur de la table la plus pleine vers la
        # table la moins pleine tant que l'écart est >= 2
        for _ in range(200):  # garde-fou anti boucle infinie
            tables = list(self.list_tables())
            if len(tables) < 2:
                break
            counts = []
            for t in tables:
                occ = self.conn.execute(
                    "SELECT COUNT(*) c FROM players WHERE table_id=? AND status='active'",
                    (t["id"],),
                ).fetchone()["c"]
                counts.append((occ, t))
            counts.sort(key=lambda x: x[0])
            smallest_count, smallest_table = counts[0]
            largest_count, largest_table = counts[-1]
            if largest_count - smallest_count < 2:
                break
            if smallest_count >= smallest_table["max_seats"]:
                break
            candidates = self.conn.execute(
                "SELECT id FROM players WHERE table_id=? AND status='active'",
                (largest_table["id"],),
            ).fetchall()
            if not candidates:
                break
            # Préfère déplacer un joueur déjà en mouvement ce rééquilibrage
            # (replacé ici suite à la fermeture d'une autre table, ou par
            # un tour précédent de cette même boucle) plutôt qu'un joueur
            # assis à cette table depuis le début : celui-ci compte déjà
            # comme "déplacé" quoi qu'il arrive, le déranger ne coûte
            # donc rien de plus — alors que déplacer quelqu'un de stable
            # sans nécessité crée un mouvement évitable dans l'historique
            # (voir onglet Mouvements). Ne change rien au résultat final
            # (nombre de joueurs par table) : seulement LEQUEL bouge.
            mover = next(
                (c for c in candidates if before_state.get(c["id"], (None, None))[0] != largest_table["id"]),
                candidates[0],
            )
            taken = {
                r["seat"]
                for r in self.conn.execute(
                    "SELECT seat FROM players WHERE table_id=? AND status='active'",
                    (smallest_table["id"],),
                )
            }
            seat = 1
            while seat in taken:
                seat += 1
            self.conn.execute(
                "UPDATE players SET table_id=?, seat=? WHERE id=?",
                (smallest_table["id"], seat, mover["id"]),
            )
            self.conn.commit()

        # NE comble PAS les sièges laissés vides par un joueur éliminé (ou
        # déplacé ailleurs) à une table par ailleurs inchangée : au poker,
        # les autres joueurs restent physiquement assis là où ils sont,
        # personne ne se déplace juste pour "combler un trou" — même
        # l'éliminateur, resté à sa place, se retrouvait pourtant avec un
        # nouveau numéro de siège à chaque élimination à sa table, sans
        # aucun mouvement annoncé (le changement de SIÈGE seul, sans
        # changement de TABLE, n'est jamais compté comme un "mouvement" —
        # voir plus bas — donc personne n'était prévenu). On ne réassigne
        # ici que les sièges qui dépassent le nombre de places actuel de la
        # table (ex : réduction du nombre de sièges par table en cours de
        # tournoi) : un vrai cas de contrainte violée, pas juste "il y a un
        # trou".
        for t in self.list_tables():
            occupants = self.conn.execute(
                "SELECT id, seat FROM players WHERE table_id=? AND status='active' "
                "ORDER BY seat",
                (t["id"],),
            ).fetchall()
            overflow = [p for p in occupants if p["seat"] > t["max_seats"]]
            if not overflow:
                continue
            taken = {p["seat"] for p in occupants if p["seat"] <= t["max_seats"]}
            seat = 1
            for p in overflow:
                while seat in taken:
                    seat += 1
                self.conn.execute("UPDATE players SET seat=? WHERE id=?", (seat, p["id"]))
                taken.add(seat)
                seat += 1
        self.conn.commit()

        # Calcule les déplacements réels (avant -> après) et les archive.
        # Seul un changement de TABLE compte comme un "mouvement" (alerte,
        # pause du chrono, historique) : un simple recompactage de numéro
        # de siège au sein de la même table (ex : combler le siège laissé
        # vide par un joueur éliminé) ne demande à personne de se déplacer
        # physiquement, donc pas d'alerte pour ça — sur un SNG à une seule
        # table, ça évite une alerte à chaque élimination alors que
        # personne ne bouge réellement de table. Les tables ne sont plus
        # jamais renommées (voir plus haut : on ferme toujours la table la
        # plus haute, jamais de renumérotation), donc un même id de table a
        # forcément le même nom avant et après — un seul dictionnaire de
        # noms suffit, plus besoin de distinguer avant/après ni de filet de
        # sécurité contre une coïncidence de nom.
        after_players = [dict(p) for p in self.list_players(status="active")]
        table_names = {t["id"]: t["name"] for t in self.list_tables(active_only=False)}
        now = time.strftime("%Y-%m-%d %H:%M:%S")
        moves = []
        for p in after_players:
            old_table_id, old_seat = before_state.get(p["id"], (None, None))
            new_table_id, new_seat = p["table_id"], p["seat"]
            if old_table_id == new_table_id:
                continue
            old_table_name = table_names.get(old_table_id)
            new_table_name = table_names.get(new_table_id)
            move = {
                "player_name": p["name"],
                "old_table_name": old_table_name,
                "old_seat": old_seat,
                "new_table_name": new_table_name,
                "new_seat": new_seat,
                "moved_at": now,
            }
            moves.append(move)
        if moves and record_moves:
            # On efface les mouvements précédents : l'onglet Mouvements
            # n'affiche que le dernier lot de déplacements en date, pas un
            # historique cumulatif.
            self.conn.execute("DELETE FROM seat_moves")
            for move in moves:
                self.conn.execute(
                    "INSERT INTO seat_moves(player_name, old_table_name, old_seat, "
                    "new_table_name, new_seat, moved_at) VALUES (?,?,?,?,?,?)",
                    (move["player_name"], move["old_table_name"], move["old_seat"],
                     move["new_table_name"], move["new_seat"], move["moved_at"]),
                )
            self.conn.commit()
        return moves

    def get_seat_moves(self, limit=500):
        """Historique des déplacements de joueurs entre tables/sièges (le
        plus récent en premier), pour l'onglet Mouvements."""
        return self.conn.execute(
            "SELECT * FROM seat_moves ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()

    def count_seat_moves(self):
        return self.conn.execute("SELECT COUNT(*) c FROM seat_moves").fetchone()["c"]

    def clear_seat_moves(self):
        """Vide l'historique des déplacements (onglet Mouvements) — utilisé
        par le bouton "Terminé" une fois les joueurs déplacés installés à
        leur nouvelle table."""
        self.conn.execute("DELETE FROM seat_moves")
        self.conn.commit()

    # ---------- primes (bounty) ----------
    def get_bounty_events(self, limit=500):
        """Historique des primes gagnées (le plus récent en premier), pour
        l'onglet Primes."""
        return self.conn.execute(
            "SELECT * FROM bounty_events ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()

    def get_active_bounties(self):
        """Joueurs actifs portant une prime, triés par prime décroissante."""
        return self.conn.execute(
            "SELECT * FROM players WHERE status='active' AND bounty > 0 "
            "ORDER BY bounty DESC"
        ).fetchall()

    def get_bounty_totals(self):
        """Classement des joueurs par cumul de primes empochées (cash),
        toutes primes gagnées, décroissant."""
        return self.conn.execute(
            "SELECT * FROM players WHERE bounty_won > 0 ORDER BY bounty_won DESC"
        ).fetchall()

    def get_presence_bonuses(self):
        """Prime de présence (en points) : chaque joueur du tournoi en
        cours reçoit `attendance_bonus_points` (réglage) pour le simple
        fait d'avoir participé à ce tournoi, 0 si le réglage est nul.
        Renvoie {nom: points}."""
        points = self.get_setting_int("attendance_bonus_points", 0)
        return {p["name"]: points for p in self.list_players()}

    def get_assiduity_bonuses(self):
        """Calcule la prime d'assiduité (en points) de chaque joueur du
        tournoi en cours. Contrôlée par deux réglages :
        - `assiduity_bonus_points` : montant en points (0/nul = prime
          désactivée) ;
        - `assiduity_consecutive_days` : nombre de présences consécutives
          requises, ce tournoi inclus (0/nul = prime désactivée aussi). 2
          = présent ce tournoi-ci ET le précédent ; 3 = ce tournoi-ci et
          les 2 précédents ; etc.
        Un joueur est éligible s'il figure dans TOUS les
        (`assiduity_consecutive_days` - 1) fichiers .tournoi précédents
        immédiats (même dossier, non récursif) — s'il manque ne serait-ce
        qu'un de ces tournois dans son historique, ou qu'il n'y a pas
        encore assez de tournois précédents, il n'est pas éligible.
        Renvoie une liste de dicts {name, present_previous, points} triée
        par nom ; liste vide si la prime est désactivée (l'un des deux
        réglages à 0)."""
        points = self.get_setting_int("assiduity_bonus_points", 0)
        consecutive_days = self.get_setting_int("assiduity_consecutive_days", 0)
        if points <= 0 or consecutive_days <= 0:
            return []

        needed_previous = consecutive_days - 1
        if needed_previous == 0:
            # 1 seule présence "consécutive" requise : ce tournoi-ci suffit.
            eligible_names = {p["name"].strip().lower() for p in self.list_players()}
        else:
            prev_files = find_previous_tournament_files(
                self.path, self.get_tournament_date(), count=needed_previous
            )
            if len(prev_files) < needed_previous:
                eligible_names = set()  # pas encore assez d'historique
            else:
                name_sets = [
                    {n.strip().lower() for n in read_player_names_from_file(fp)}
                    for fp in prev_files
                ]
                eligible_names = set.intersection(*name_sets)

        result = [
            {
                "name": p["name"],
                "present_previous": p["name"].strip().lower() in eligible_names,
                "points": points if p["name"].strip().lower() in eligible_names else 0,
            }
            for p in self.list_players()
        ]
        result.sort(key=lambda r: r["name"].casefold())
        return result

    def get_ranking_bonuses(self):
        """Calcule la prime de classement (en points) de chaque joueur dont
        le rang est déjà connu : un joueur éliminé (place déjà attribuée),
        ou le vainqueur une fois le tournoi terminé (même convention que le
        rang affiché dans l'onglet Joueurs / les exports). Les joueurs
        encore actifs en cours de tournoi (rang pas encore connu) n'ont pas
        de ligne. Renvoie une liste de dicts {name, place, nombre, valeur,
        montant} (montant = nombre × valeur), triée par rang croissant."""
        flat_value = self.get_setting_int("ranking_bonus_points", 0)
        n_players = self.get_stats()["total_players_ever"]
        all_players = self.list_players()
        active = [p for p in all_players if p["status"] == "active"]
        finished = len(active) == 1

        result = []
        for p in all_players:
            place = None
            if p["status"] == "eliminated":
                place = p["place"]
            elif p["status"] == "active" and finished:
                place = 1
            if place is None:
                continue
            valeur = ranking_points(place, n_players, flat_value)
            result.append({
                "name": p["name"], "place": place,
                "nombre": 1, "valeur": valeur, "montant": valeur,
            })
        result.sort(key=lambda r: r["place"])
        return result

    def get_bounty_bonuses(self):
        """Calcule la prime de bounty (en points) de chaque joueur du
        tournoi en cours : Nombre = nombre de joueurs qu'il a éliminés
        (players.kills, incrémenté sur toute élimination avec éliminateur
        désigné — indépendant de l'ancien mécanisme de bounty en €/PKO),
        Valeur = réglage manuel `bounty_amount` s'il est non nul, sinon
        10×√N points par bounty (N = nombre total de joueurs du tournoi),
        Montant = Nombre × Valeur. Renvoie une liste de dicts
        {name, nombre, valeur, montant} pour tous les joueurs, triée par
        montant décroissant."""
        flat_value = self.get_setting_int("bounty_amount", 0)
        n_players = self.get_stats()["total_players_ever"]
        valeur = bounty_unit_value(n_players, flat_value)
        result = [
            {
                "name": p["name"], "nombre": p["kills"],
                "valeur": valeur, "montant": p["kills"] * valeur,
            }
            for p in self.list_players()
        ]
        result.sort(key=lambda r: (-r["montant"], r["name"].casefold()))
        return result

    def get_primes_summary(self, sort_column=None, ascending=True):
        """Construit, pour chaque joueur du tournoi en cours, la ligne
        récapitulative des primes en points affichée dans l'onglet Primes
        (présence, assiduité, rang, classement, bounty nombre/valeur/
        montant, TOTAL). Utilisé aussi pour son export dédié.

        `sort_column` : 'rang', 'bo_nombre' ou 'total' (autre valeur ou
        None -> tri par défaut, TOTAL décroissant). Les valeurs manquantes
        (ex : rang d'un joueur encore actif) sont toujours reléguées en
        fin de liste, quel que soit le sens du tri."""
        presence_by_name = self.get_presence_bonuses()
        assiduity_by_name = {r["name"]: r for r in self.get_assiduity_bonuses()}
        ranking_by_name = {r["name"]: r for r in self.get_ranking_bonuses()}
        bounty_by_name = {r["name"]: r for r in self.get_bounty_bonuses()}

        rows = []
        for p in self.list_players():
            name = p["name"]
            presence = presence_by_name.get(name, 0)
            assiduite = assiduity_by_name.get(name, {}).get("points", 0)
            rk = ranking_by_name.get(name)
            rang, cl_montant = (rk["place"], rk["montant"]) if rk else (None, 0)
            bt = bounty_by_name.get(name, {"nombre": 0, "valeur": 0, "montant": 0})
            total = presence + assiduite + cl_montant + bt["montant"]
            rows.append({
                "name": name, "presence": presence, "assiduite": assiduite,
                "rang": rang, "cl_montant": cl_montant,
                "bo_nombre": bt["nombre"], "bo_valeur": bt["valeur"], "bo_montant": bt["montant"],
                "total": total,
            })

        if sort_column in ("rang", "bo_nombre", "total"):
            def sort_key(r):
                v = r[sort_column]
                if v is None:
                    return (1, 0, r["name"].casefold())
                return (0, v if ascending else -v, r["name"].casefold())
            rows.sort(key=sort_key)
        else:
            rows.sort(key=lambda r: (-r["total"], r["name"].casefold()))
        return rows

    def export_primes_csv(self, path, columns=None, sort_column=None, ascending=True):
        """Exporte le tableau de l'onglet Primes tel qu'affiché, en CSV.
        `columns` : sous-ensemble de clés de PRIMES_COLUMNS (None =
        toutes). `sort_column`/`ascending` : voir get_primes_summary."""
        import csv

        cols = _selected_period_columns(PRIMES_COLUMNS, columns)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([h for _, h, _ in cols])
            for r in self.get_primes_summary(sort_column=sort_column, ascending=ascending):
                writer.writerow([fn(r) for _, _, fn in cols])
        return path

    def export_primes_xlsx(self, path, columns=None, sort_column=None, ascending=True, title=None):
        """Exporte le tableau de l'onglet Primes au format Excel (.xlsx).
        `columns`, `sort_column`, `ascending` : voir export_primes_csv.
        `title` : remplace le titre par défaut (nom du tournoi) si fourni.
        Nécessite 'openpyxl'."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        cols = _selected_period_columns(PRIMES_COLUMNS, columns)
        rows = self.get_primes_summary(sort_column=sort_column, ascending=ascending)

        wb = Workbook()
        ws = wb.active
        ws.title = "Primes"

        name = title or self.get_setting("tournament_name", "Tournoi")
        ws.append([name])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append(["Primes en points : présence, assiduité, classement, bounty"])
        ws["A2"].font = Font(italic=True)
        ws.append([])

        headers = [h for _, h, _ in cols]
        header_row = ws.max_row + 1
        ws.append(headers)
        header_fill = PatternFill(start_color="1F4E24", end_color="1F4E24", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            ws.append([fn(r) for _, _, fn in cols])

        widths = {"name": 22, "presence": 12, "assiduite": 12, "rang": 10,
                  "cl_montant": 14, "bo_nombre": 12, "bo_valeur": 12,
                  "bo_montant": 12, "total": 12}
        for i, (key, _, _) in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 14)

        wb.save(path)
        return path

    def export_movement_slips_pdf(self, path, n_slips=6):
        """Génère une page de coupons VIERGES (rien de préempli, pas même
        le nom du tournoi) à découper, un par joueur concerné par un
        changement de table — bouton "Imprimer" de l'onglet Mouvements.
        Le responsable écrit chaque coupon à la main au moment du
        mouvement réel (nom du tournoi, nom du joueur, ancienne table/
        siège, nouvelle table/siège) et le remet directement au joueur :
        plus rapide et plus discret que d'annoncer les mouvements à voix
        haute. `n_slips` : nombre de coupons identiques sur la page (une
        seule impression sert pour tout un mouvement de plusieurs
        joueurs à la fois). Nécessite 'fpdf2'."""
        from fpdf import FPDF

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=False)
        pdf.add_page()

        margin = 12
        usable_width = pdf.w - 2 * margin
        slip_height = (pdf.h - 2 * margin) / n_slips
        line_h = (slip_height - 8) / 4
        half_width = (usable_width - 6) / 2

        pdf.set_font("Helvetica", "", 11)
        for i in range(n_slips):
            top = margin + i * slip_height
            if i > 0:
                pdf.dashed_line(margin, top, pdf.w - margin, top, dash_length=2, space_length=1.5)
            y = top + 5

            def field(label, x, y, w):
                pdf.set_xy(x, y)
                pdf.cell(w, line_h, _pdf_text(label), border="B")

            field("Nom du tournoi : ", margin, y, usable_width)
            field("Nom du joueur : ", margin, y + line_h, usable_width)
            field("Ancienne table : ", margin, y + 2 * line_h, half_width)
            field("Ancien siège : ", margin + half_width + 6, y + 2 * line_h, half_width)
            field("Nouvelle table : ", margin, y + 3 * line_h, half_width)
            field("Nouveau siège : ", margin + half_width + 6, y + 3 * line_h, half_width)

        pdf.output(path)
        return path

    def export_movement_slips_filled_pdf(self, path, n_per_page=6):
        """Comme export_movement_slips_pdf, mais un coupon par mouvement
        RÉELLEMENT en attente (voir get_seat_moves, même liste que le
        tableau de l'onglet Mouvements), déjà rempli avec les vraies
        valeurs — rien à écrire à la main, juste à découper et remettre.
        Pratique quand beaucoup de joueurs sont concernés à la fois (voir
        bouton "Imprimer" de l'onglet Mouvements, à distinguer du bouton
        "Imprimer Vierge" qui imprime des coupons vides). Une page par
        tranche de `n_per_page` mouvements. Lève ValueError si aucun
        mouvement n'est en attente. Nécessite 'fpdf2'."""
        moves = self.get_seat_moves()
        if not moves:
            raise ValueError("Aucun mouvement en attente à imprimer.")
        name = self.get_setting("tournament_name", "Tournoi")

        from fpdf import FPDF

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=False)

        margin = 12
        usable_width = pdf.w - 2 * margin
        slip_height = (pdf.h - 2 * margin) / n_per_page
        line_h = (slip_height - 8) / 4
        half_width = (usable_width - 6) / 2

        def field(label, value, x, y, w):
            pdf.set_xy(x, y)
            pdf.cell(w, line_h, _pdf_text(f"{label}{value}"), border="B")

        for i, m in enumerate(moves):
            pos = i % n_per_page
            if pos == 0:
                pdf.add_page()
                pdf.set_font("Helvetica", "", 11)
            top = margin + pos * slip_height
            if pos > 0:
                pdf.dashed_line(margin, top, pdf.w - margin, top, dash_length=2, space_length=1.5)
            y = top + 5
            field("Nom du tournoi : ", name, margin, y, usable_width)
            field("Nom du joueur : ", m["player_name"], margin, y + line_h, usable_width)
            field("Ancienne table : ", m["old_table_name"] or "-", margin, y + 2 * line_h, half_width)
            field("Ancien siège : ", m["old_seat"] or "-", margin + half_width + 6, y + 2 * line_h, half_width)
            field("Nouvelle table : ", m["new_table_name"] or "-", margin, y + 3 * line_h, half_width)
            field("Nouveau siège : ", m["new_seat"] or "-", margin + half_width + 6, y + 3 * line_h, half_width)

        pdf.output(path)
        return path

    def export_primes_pdf(self, path, columns=None, sort_column=None, ascending=True, title=None):
        """Exporte le tableau de l'onglet Primes en PDF. `columns`,
        `sort_column`, `ascending` : voir export_primes_csv. `title` : voir
        export_primes_xlsx. Nécessite 'fpdf2'."""
        cols = _selected_period_columns(PRIMES_COLUMNS, columns)
        rows = self.get_primes_summary(sort_column=sort_column, ascending=ascending)
        name = title or self.get_setting("tournament_name", "Tournoi")
        return _write_pdf_table(
            path, name,
            ["Primes en points : présence, assiduité, classement, bounty"],
            [h for _, h, _ in cols],
            [[fn(r) for _, _, fn in cols] for r in rows],
        )

    def export_settings_pdf(self, path, club_name=None):
        """Exporte en PDF tous les réglages actuels de ce tournoi (onglet
        Paramètres, voir SETTINGS_PRINT_FIELDS), sous forme d'un tableau
        Réglage/Valeur — pour en garder une trace papier ou la partager.
        Reflète les valeurs déjà enregistrées dans ce fichier .tournoi
        (l'appelant, App._print_settings_pdf, enregistre d'abord le
        formulaire pour être sûr qu'elles soient à jour). Booléens
        (pko_mode) affichés "Oui"/"Non" plutôt que "1"/"0". `club_name` :
        à passer explicitement par l'appelant (ex : export_prefs.
        load_value("club_name", "")) — ce réglage, commun à tous les
        tournois/Sit & Go, n'est justement JAMAIS écrit dans ce fichier
        .tournoi (voir App._collect_and_save_all_settings), donc
        self.get_setting("club_name", ...) renverrait toujours vide ici.
        Nécessite 'fpdf2'."""
        name = self.get_setting("tournament_name", "Tournoi")
        rows = []
        for key, label in SETTINGS_PRINT_FIELDS:
            if key == "club_name":
                value = club_name or ""
            else:
                value = self.get_setting(key, "")
                if key == "pko_mode":
                    value = "Oui" if value in ("1", 1, True) else "Non"
            rows.append((label, value))
        return _write_pdf_table(
            path, f"Paramètres — {name}",
            [f"Imprimé le {time.strftime('%d/%m/%Y %H:%M')}"],
            ["Réglage", "Valeur"],
            rows,
        )

    def export_bounty_history_csv(self, path, columns=None):
        """Exporte l'historique du bounty progressif (mécanisme PKO
        interne, 2e tableau de l'onglet Primes), en CSV. `columns` :
        sous-ensemble de clés de BOUNTY_HISTORY_COLUMNS (None = toutes)."""
        import csv

        cols = _selected_period_columns(BOUNTY_HISTORY_COLUMNS, columns)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([h for _, h, _ in cols])
            for r in self.get_bounty_events(limit=10000):
                writer.writerow([fn(r) for _, _, fn in cols])
        return path

    def export_bounty_history_xlsx(self, path, columns=None, title=None):
        """Exporte l'historique du bounty progressif au format Excel
        (.xlsx). `columns` : voir export_bounty_history_csv. `title` :
        remplace le titre par défaut (nom du tournoi) si fourni. Nécessite
        'openpyxl'."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        cols = _selected_period_columns(BOUNTY_HISTORY_COLUMNS, columns)
        rows = self.get_bounty_events(limit=10000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Historique bounty"

        name = title or self.get_setting("tournament_name", "Tournoi")
        ws.append([name])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append(["Historique du bounty progressif (mécanisme PKO interne)"])
        ws["A2"].font = Font(italic=True)
        ws.append([])

        headers = [h for _, h, _ in cols]
        header_row = ws.max_row + 1
        ws.append(headers)
        header_fill = PatternFill(start_color="1F4E24", end_color="1F4E24", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            ws.append([fn(r) for _, _, fn in cols])

        widths = {"time": 18, "eliminated": 20, "eliminator": 20, "amount": 14, "grow": 16}
        for i, (key, _, _) in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 14)

        wb.save(path)
        return path

    def export_bounty_history_pdf(self, path, columns=None, title=None):
        """Exporte l'historique du bounty progressif en PDF. `columns` :
        voir export_bounty_history_csv. `title` : voir
        export_bounty_history_xlsx. Nécessite 'fpdf2'."""
        cols = _selected_period_columns(BOUNTY_HISTORY_COLUMNS, columns)
        rows = self.get_bounty_events(limit=10000)
        name = title or self.get_setting("tournament_name", "Tournoi")
        return _write_pdf_table(
            path, name,
            ["Historique du bounty progressif (mécanisme PKO interne)"],
            [h for _, h, _ in cols],
            [[fn(r) for _, _, fn in cols] for r in rows],
        )

    # ---------- blind structure ----------
    def get_blind_structure(self):
        return self.conn.execute(
            "SELECT * FROM blind_levels ORDER BY level_order"
        ).fetchall()

    def set_blind_structure(self, levels):
        """levels: liste de dicts {small_blind, big_blind, ante, duration_minutes,
        is_break, break_label}"""
        self.conn.execute("DELETE FROM blind_levels")
        for i, lvl in enumerate(levels, start=1):
            self.conn.execute(
                "INSERT INTO blind_levels(level_order, small_blind, big_blind, "
                "ante, duration_minutes, is_break, break_label) VALUES (?,?,?,?,?,?,?)",
                (
                    i,
                    lvl.get("small_blind", 0),
                    lvl.get("big_blind", 0),
                    lvl.get("ante", 0),
                    lvl.get("duration_minutes", 15),
                    1 if lvl.get("is_break") else 0,
                    lvl.get("break_label", "Pause"),
                ),
            )
        self.conn.commit()

    def get_current_level(self):
        order = self.get_setting_int("current_level_order", 1)
        row = self.conn.execute(
            "SELECT * FROM blind_levels WHERE level_order=?", (order,)
        ).fetchone()
        if row is None:
            row = self.conn.execute(
                "SELECT * FROM blind_levels ORDER BY level_order LIMIT 1"
            ).fetchone()
        return row

    def get_next_level(self):
        order = self.get_setting_int("current_level_order", 1)
        return self.conn.execute(
            "SELECT * FROM blind_levels WHERE level_order=?", (order + 1,)
        ).fetchone()

    def get_round_number(self, level_order):
        """Numéro de round au sens de l'onglet Blindes pour la ligne
        level_order donnée (une pause n'y compte pas comme un round à
        part entière, contrairement à level_order qui numérote toutes
        les lignes de la structure, pauses comprises — voir "Niveau" au
        Chronomètre/écran projecteur, qui utilisait jusqu'ici directement
        level_order et pouvait donc afficher un numéro différent de la
        colonne "Round" de l'onglet Blindes dès qu'une pause avait eu
        lieu). Renvoie None si level_order est vide/invalide."""
        if not level_order:
            return None
        row = self.conn.execute(
            "SELECT COUNT(*) c FROM blind_levels WHERE is_break=0 AND level_order<=?",
            (level_order,),
        ).fetchone()
        return row["c"] or None

    def get_current_round_number(self):
        """Numéro de round (voir get_round_number) du niveau actuellement
        en cours — utilisé pour horodater les éliminations (colonne
        "Round" de l'onglet Joueurs)."""
        return self.get_round_number(self.get_setting_int("current_level_order", 0))

    # ---------- payout structure ----------
    def get_payout_structure(self):
        return self.conn.execute(
            "SELECT * FROM payout_structure ORDER BY place"
        ).fetchall()

    def set_payout_structure(self, place_to_pct):
        self.conn.execute("DELETE FROM payout_structure")
        for place, pct in place_to_pct.items():
            self.conn.execute(
                "INSERT INTO payout_structure(place, percentage) VALUES (?, ?)",
                (place, pct),
            )
        self.conn.commit()

    # ---------- stats / prize pool ----------
    def get_stats(self):
        all_players = self.list_players()
        active = [p for p in all_players if p["status"] == "active"]
        entries = sum(p["buyin_count"] for p in all_players)
        rebuys = sum(p["rebuy_count"] for p in all_players)
        addons = sum(p["addon_count"] for p in all_players)
        buyin_amount = self.get_setting_float("buyin_amount", 0)
        rebuy_amount = self.get_setting_float("rebuy_amount", 0)
        addon_amount = self.get_setting_float("addon_amount", 0)
        rake_pct = self.get_setting_float("rake_percent", 0)
        gross = entries * buyin_amount + rebuys * rebuy_amount + addons * addon_amount
        prize_pool = gross * (1 - rake_pct / 100.0)
        # Tapis moyen = total des chips en jeu / joueurs encore actifs. Le
        # total doit porter sur les actifs ET les éliminés (mais pas les
        # forfaits — voir plus bas) : les chips d'un joueur éliminé ne
        # disparaissent pas de la table, elles passent dans le tapis de
        # celui qui l'a éliminé — mais ce transfert n'est pas forcément
        # ressaisi manuellement (onglet Joueurs > Modifier les chips)
        # pour chaque main. Se limiter aux actifs ferait baisser le total
        # à chaque élimination et donnerait un tapis moyen sous-évalué.
        # Les forfaits (status='withdrawn') sont exclus : leurs chips ne
        # sont jamais entrées en jeu / en sont sorties avec eux, elles ne
        # doivent pas gonfler le total.
        total_chips = sum(p["chips"] for p in all_players if p["status"] != "withdrawn")
        avg_stack = total_chips / len(active) if active else 0

        # Durée du tournoi (temps réel écoulé depuis le tout premier
        # "Démarrer" — voir App._clock_resume) : continue de courir tant
        # que la partie n'est pas terminée (y compris pendant les pauses
        # du chrono de niveau, qui n'arrêtent pas le temps réel), puis se
        # fige à l'heure de fin dès qu'il ne reste plus qu'1 joueur actif
        # (voir eliminate_player) plutôt que de continuer à défiler alors
        # que tout le monde est déjà parti.
        start_epoch = self.get_setting_int("tournament_start_epoch", 0)
        if start_epoch == 0 and self.get_setting_int("clock_started", 0) == 1:
            # Tournoi déjà en cours avant l'ajout de ce réglage (le chrono
            # avait déjà été démarré) : l'heure du tout premier "Démarrer"
            # n'a jamais été enregistrée, donc "Durée" resterait bloquée à
            # 00:00:00 pour toujours sans ce rattrapage ponctuel. On
            # l'approxime une bonne fois pour toutes avec l'heure de la
            # première élimination déjà enregistrée si elle existe
            # (meilleure estimation disponible), sinon avec l'heure
            # actuelle — puis on la fige en réglage pour ne plus jamais y
            # revenir (sans quoi la durée repartirait de zéro à chaque
            # rafraîchissement).
            earliest = None
            for p in all_players:
                if p["elim_time"]:
                    try:
                        t = time.mktime(time.strptime(p["elim_time"], "%Y-%m-%d %H:%M:%S"))
                    except ValueError:
                        continue
                    if earliest is None or t < earliest:
                        earliest = t
            start_epoch = int(earliest) if earliest else int(time.time())
            self.set_setting("tournament_start_epoch", start_epoch)
        end_epoch = self.get_setting_int("tournament_end_epoch", 0)
        if start_epoch:
            duration_seconds = max(0, (end_epoch or int(time.time())) - start_epoch)
        else:
            duration_seconds = 0

        return {
            "total_players_ever": len(all_players),
            "active_count": len(active),
            "entries": entries,
            "rebuys": rebuys,
            "addons": addons,
            "gross": gross,
            "prize_pool": prize_pool,
            "total_chips": total_chips,
            "avg_stack": avg_stack,
            "duration_seconds": duration_seconds,
            "tournament_finished": end_epoch != 0,
        }

    def get_live_status(self):
        """Résumé de l'état courant du tournoi, pour l'affichage dans le
        Lobby SNG (liste de plusieurs tournois à la fois) : nom, date,
        joueurs actifs/total, niveau de blindes courant (ou pause), temps
        restant dans ce niveau, chrono démarré/en pause, tournoi terminé.
        Ne modifie rien (ne fait pas avancer automatiquement de niveau,
        contrairement à l'onglet Chronomètre — une simple consultation ne
        doit pas altérer le déroulé du tournoi)."""
        stats = self.get_stats()
        level = self.get_current_level()
        clock_started = self.get_setting_int("clock_started", 0) == 1
        is_paused = self.get_setting_int("is_paused", 1) == 1
        remaining_seconds = None
        if level is not None:
            duration = level["duration_minutes"] * 60
            if not clock_started:
                elapsed = 0
            elif is_paused:
                elapsed = self.get_setting_int("paused_accum_seconds", 0)
            else:
                start = self.get_setting_int("level_start_epoch", int(time.time()))
                elapsed = int(time.time()) - start
            remaining_seconds = max(0, duration - elapsed)
        finished = stats["active_count"] <= 1 and stats["total_players_ever"] > 1
        return {
            "name": self.get_setting("tournament_name", "Tournoi"),
            "date": self.get_tournament_date(),
            "active_count": stats["active_count"],
            "total_players_ever": stats["total_players_ever"],
            "level": level,
            "remaining_seconds": remaining_seconds,
            "clock_started": clock_started,
            "is_paused": is_paused,
            "finished": finished,
        }

    def get_payouts_amounts(self):
        stats = self.get_stats()
        pool = stats["prize_pool"]
        result = []
        for row in self.get_payout_structure():
            result.append(
                {
                    "place": row["place"],
                    "percentage": row["percentage"],
                    "amount": pool * row["percentage"] / 100.0,
                }
            )
        return result

    def _results_rows(self):
        """Construit les lignes de résultats (rang, nom, statut, gain,
        buy-ins, rebuys, add-ons, prime gagnée) sous forme de dicts, triées
        meilleur rang en premier. Utilisé par les exports CSV et XLSX.

        Le rang suit exactement la même convention que l'onglet Joueurs :
        1 pour le vainqueur (seul joueur encore actif, tournoi terminé),
        None (affiché "-") pour les autres joueurs encore actifs tant que
        le tournoi est en cours, et le rang habituel pour les éliminés."""
        payouts_by_place = {p["place"]: p["amount"] for p in self.get_payouts_amounts()}
        eliminated = [p for p in self.list_players() if p["status"] == "eliminated"]
        eliminated.sort(key=lambda p: p["place"])
        withdrawn = [p for p in self.list_players() if p["status"] == "withdrawn"]
        active = [p for p in self.list_players() if p["status"] == "active"]
        finished = len(active) == 1

        rows = []
        for p in active:
            rows.append({
                "rang": 1 if finished else None,
                "name": p["name"], "status": "En cours" if not finished else "Terminé",
                "gain": payouts_by_place.get(1) if finished else None,
                "buyin": p["buyin_count"], "rebuy": p["rebuy_count"], "addon": p["addon_count"],
                "bounty_won": p["bounty_won"],
            })
        for p in eliminated:
            rows.append({
                "rang": p["place"],
                "name": p["name"], "status": "Éliminé",
                "gain": payouts_by_place.get(p["place"]),
                "buyin": p["buyin_count"], "rebuy": p["rebuy_count"], "addon": p["addon_count"],
                "bounty_won": p["bounty_won"],
            })
        for p in withdrawn:
            rows.append({
                "rang": None,
                "name": p["name"], "status": "Forfait", "gain": None,
                "buyin": p["buyin_count"], "rebuy": p["rebuy_count"], "addon": p["addon_count"],
                "bounty_won": p["bounty_won"],
            })
        return rows

    def players_rows(self, sort_column=None, ascending=True):
        """Construit les lignes du tableau de l'onglet Joueurs (nom, table,
        siège, chips, achats, prime en jeu, statut, rang), avec le même
        calcul de rang que cet onglet. Utilisé pour son export dédié.

        `sort_column`/`ascending` reproduisent exactement le tri appliqué
        dans l'onglet (voir App._sort_players_by côté interface) : sans
        eux, l'ordre d'affichage à l'écran (par ex. trié par Rang) et
        l'ordre du fichier exporté pouvaient diverger."""
        status_labels = {"active": "Actif", "withdrawn": "Forfait", "eliminated": "Éliminé"}
        tables = {t["id"]: t["name"] for t in self.list_tables(active_only=False)}
        players = [dict(p) for p in self.list_players()]
        n_active = sum(1 for p in players if p["status"] == "active")

        rows = []
        for p in players:
            if p["status"] == "active":
                rang = 1 if n_active == 1 else None
            elif p["status"] == "eliminated":
                rang = p["place"]
            else:
                rang = None
            rows.append({
                "name": p["name"],
                "club": p["club"] or "",
                "table": tables.get(p["table_id"], "-") if p["table_id"] else "-",
                "seat": p["seat"],
                "chips": p["chips"],
                "buyin": p["buyin_count"],
                "rebuy": p["rebuy_count"],
                "addon": p["addon_count"],
                "bounty": p["bounty"],
                "status": status_labels.get(p["status"], p["status"]),
                "rang": rang,
                "elim_time": p["elim_time"] or "",
                "elim_round": p["elim_round"],
                "eliminated_by": p["eliminated_by_name"] or "",
            })

        if sort_column == "name":
            rows.sort(key=lambda r: r["name"].lower())
        elif sort_column == "status":
            rows.sort(key=lambda r: r["status"].lower())
        elif sort_column == "table":
            rows.sort(key=lambda r: ((r["table"] or "").lower(), r["seat"] or 0))
        elif sort_column == "rang":
            rows.sort(key=lambda r: r["rang"] or 1)
        elif sort_column == "elim_time":
            rows.sort(key=lambda r: r["elim_time"] or "")
        elif sort_column == "eliminated_by":
            rows.sort(key=lambda r: r["eliminated_by"].lower())
        if sort_column and not ascending:
            rows.reverse()
        return rows

    def _bounty_in_use(self):
        if self.get_setting_int("bounty_amount", 0) > 0:
            return True
        return self.conn.execute(
            "SELECT COUNT(*) c FROM players WHERE bounty_won > 0"
        ).fetchone()["c"] > 0

    def export_results_csv(self, path, columns=None):
        """Exporte le classement final (éliminés triés par rang, puis
        joueurs encore actifs) avec le gain correspondant, en CSV.
        `columns` : sous-ensemble de clés de RESULT_COLUMNS à inclure
        (None = toutes celles pertinentes, primes comprises seulement si
        le bounty est utilisé)."""
        import csv

        cols = _selected_period_columns(RESULT_COLUMNS, columns)
        if columns is None and not self._bounty_in_use():
            cols = [c for c in cols if c[0] != "bounty_won"]
        rows = self._results_rows()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([h for _, h, _ in cols])
            for r in rows:
                writer.writerow([fn(r) for _, _, fn in cols])
        return path

    def export_results_xlsx(self, path, columns=None):
        """Exporte le classement final au format Excel (.xlsx), avec
        mise en forme (en-têtes en gras, colonnes ajustées, ligne de
        synthèse du tournoi). `columns` : voir export_results_csv.
        Nécessite le paquet 'openpyxl'."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        cols = _selected_period_columns(RESULT_COLUMNS, columns)
        if columns is None and not self._bounty_in_use():
            cols = [c for c in cols if c[0] != "bounty_won"]
        rows = self._results_rows()

        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"

        stats = self.get_stats()
        name = self.get_setting("tournament_name", "Tournoi")

        ws.append([name])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"Entrées : {stats['entries']}    Prize pool : {stats['prize_pool']:.2f} €"])
        ws["A2"].font = Font(italic=True)
        ws.append([])

        headers = [h for _, h, _ in cols]
        header_row = ws.max_row + 1
        ws.append(headers)
        header_fill = PatternFill(start_color="1F4E24", end_color="1F4E24", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            ws.append([fn(r) for _, _, fn in cols])

        widths = {"rang": 10, "name": 26, "status": 12, "gain": 14,
                  "buyin": 10, "rebuy": 10, "addon": 10, "bounty_won": 16}
        for i, (key, _, _) in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 14)

        wb.save(path)
        return path

    def export_results_pdf(self, path, columns=None):
        """Exporte le classement final en PDF. `columns` : voir
        export_results_csv. Nécessite le paquet 'fpdf2'."""
        cols = _selected_period_columns(RESULT_COLUMNS, columns)
        if columns is None and not self._bounty_in_use():
            cols = [c for c in cols if c[0] != "bounty_won"]
        rows = self._results_rows()
        stats = self.get_stats()
        name = self.get_setting("tournament_name", "Tournoi")
        return _write_pdf_table(
            path, name,
            [f"Entrées : {stats['entries']}    Prize pool : {stats['prize_pool']:.2f} EUR"],
            [h for _, h, _ in cols],
            [[fn(r) for _, _, fn in cols] for r in rows],
        )

    def export_payouts_csv(self, path, columns=None):
        """Exporte la grille de gains telle qu'affichée dans l'onglet
        Gains (place, pourcentage, montant — sans nom de joueur), en CSV.
        `columns` : sous-ensemble de clés de PAYOUT_COLUMNS (None = toutes)."""
        import csv

        cols = _selected_period_columns(PAYOUT_COLUMNS, columns)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([h for _, h, _ in cols])
            for r in self.get_payouts_amounts():
                writer.writerow([fn(r) for _, _, fn in cols])
        return path

    def export_payouts_xlsx(self, path, columns=None):
        """Exporte la grille de gains au format Excel (.xlsx). `columns` :
        voir export_payouts_csv. Nécessite le paquet 'openpyxl'."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        cols = _selected_period_columns(PAYOUT_COLUMNS, columns)
        rows = self.get_payouts_amounts()

        wb = Workbook()
        ws = wb.active
        ws.title = "Grille de gains"

        stats = self.get_stats()
        name = self.get_setting("tournament_name", "Tournoi")

        ws.append([name])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"Entrées : {stats['entries']}    Prize pool : {stats['prize_pool']:.2f} €"])
        ws["A2"].font = Font(italic=True)
        ws.append([])

        headers = [h for _, h, _ in cols]
        header_row = ws.max_row + 1
        ws.append(headers)
        header_fill = PatternFill(start_color="1F4E24", end_color="1F4E24", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            ws.append([fn(r) for _, _, fn in cols])

        widths = {"place": 10, "percentage": 16, "amount": 14}
        for i, (key, _, _) in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 14)

        wb.save(path)
        return path

    def export_payouts_pdf(self, path, columns=None):
        """Exporte la grille de gains en PDF. `columns` : voir
        export_payouts_csv. Nécessite le paquet 'fpdf2'."""
        cols = _selected_period_columns(PAYOUT_COLUMNS, columns)
        rows = self.get_payouts_amounts()
        stats = self.get_stats()
        name = self.get_setting("tournament_name", "Tournoi")
        return _write_pdf_table(
            path, name,
            [f"Entrées : {stats['entries']}    Prize pool : {stats['prize_pool']:.2f} EUR"],
            [h for _, h, _ in cols],
            [[fn(r) for _, _, fn in cols] for r in rows],
        )

    def export_players_csv(self, path, columns=None, sort_column=None, ascending=True):
        """Exporte le tableau de l'onglet Joueurs tel qu'affiché (nom,
        table, siège, chips, achats, prime en jeu, statut, rang), en CSV.
        `columns` : sous-ensemble de clés de PLAYERS_TAB_COLUMNS (None =
        toutes). `sort_column`/`ascending` : voir players_rows — reprend
        le tri actuellement appliqué dans l'onglet."""
        import csv

        cols = _selected_period_columns(PLAYERS_TAB_COLUMNS, columns)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([h for _, h, _ in cols])
            for r in self.players_rows(sort_column=sort_column, ascending=ascending):
                writer.writerow([fn(r) for _, _, fn in cols])
        return path

    def export_players_xlsx(self, path, columns=None, sort_column=None, ascending=True,
                             title=None, show_prize_pool=True):
        """Exporte le tableau de l'onglet Joueurs au format Excel (.xlsx).
        `columns`, `sort_column`, `ascending` : voir export_players_csv.
        `title` : remplace le titre par défaut (nom du tournoi) si fourni —
        utilisé par exemple par l'export du Classement. `show_prize_pool` :
        si False, n'affiche pas la ligne "Entrées : ... Prize pool : ..."
        (également utilisé par l'export du Classement, où ça n'a pas de
        sens). Nécessite 'openpyxl'."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        cols = _selected_period_columns(PLAYERS_TAB_COLUMNS, columns)
        rows = self.players_rows(sort_column=sort_column, ascending=ascending)

        wb = Workbook()
        ws = wb.active
        ws.title = "Joueurs"

        stats = self.get_stats()
        name = title or self.get_setting("tournament_name", "Tournoi")

        ws.append([name])
        ws["A1"].font = Font(bold=True, size=14)
        if show_prize_pool:
            ws.append([f"Entrées : {stats['entries']}    Prize pool : {stats['prize_pool']:.2f} €"])
            ws[f"A{ws.max_row}"].font = Font(italic=True)
        ws.append([])

        headers = [h for _, h, _ in cols]
        header_row = ws.max_row + 1
        ws.append(headers)
        header_fill = PatternFill(start_color="1F4E24", end_color="1F4E24", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            ws.append([fn(r) for _, _, fn in cols])

        widths = {"name": 22, "table": 14, "seat": 10, "chips": 12,
                  "buyin": 10, "rebuy": 10, "addon": 10, "bounty": 12,
                  "status": 12, "rang": 10, "elim_time": 18, "elim_round": 10,
                  "eliminated_by": 18}
        for i, (key, _, _) in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 14)

        wb.save(path)
        return path

    def export_players_pdf(self, path, columns=None, sort_column=None, ascending=True,
                            title=None, show_prize_pool=True):
        """Exporte le tableau de l'onglet Joueurs en PDF. `columns`,
        `sort_column`, `ascending` : voir export_players_csv. `title`,
        `show_prize_pool` : voir export_players_xlsx. Nécessite le paquet
        'fpdf2'."""
        cols = _selected_period_columns(PLAYERS_TAB_COLUMNS, columns)
        rows = self.players_rows(sort_column=sort_column, ascending=ascending)
        stats = self.get_stats()
        name = title or self.get_setting("tournament_name", "Tournoi")
        subtitle_lines = [f"Entrées : {stats['entries']}    Prize pool : {stats['prize_pool']:.2f} EUR"] if show_prize_pool else []
        return _write_pdf_table(
            path, name,
            subtitle_lines,
            [h for _, h, _ in cols],
            [[fn(r) for _, _, fn in cols] for r in rows],
        )

    def close(self):
        self.conn.close()


def read_player_names_from_file(path):
    """Lit uniquement les noms des joueurs d'un fichier .tournoi existant,
    triés par ordre alphabétique, sans toucher au fichier ni reprendre
    leurs performances (chips, place, buy-ins...). Utilisé pour reprendre
    la liste des joueurs d'un tournoi précédent dans un nouveau tournoi.
    Ouvre la base en lecture seule (URI mode=ro) pour ne jamais créer ni
    modifier ce fichier, même par erreur. Lève une exception si le
    fichier n'est pas une base de tournoi valide."""
    uri = f"file:{os.path.abspath(path)}?mode=ro"
    conn = sqlite3.connect(uri, uri=True)
    try:
        rows = conn.execute(
            "SELECT DISTINCT name FROM players ORDER BY name COLLATE NOCASE"
        ).fetchall()
        return [row[0] for row in rows if row[0] and row[0].strip()]
    finally:
        conn.close()


# =====================================================================
# Synthèse multi-tournois par période (parcourt plusieurs fichiers
# .tournoi d'un dossier). Contrairement au reste de ce module, ces
# fonctions ne portent pas sur un seul tournoi mais en agrègent
# plusieurs — elles sont donc au niveau module plutôt que sur la classe
# Database.
# =====================================================================

def find_tournament_files(folder, recursive=True):
    """Liste (triée) des fichiers .tournoi trouvés dans `folder`, et ses
    sous-dossiers si `recursive` est vrai."""
    pattern = os.path.join(folder, "**", "*.tournoi") if recursive else os.path.join(folder, "*.tournoi")
    return sorted(glob.glob(pattern, recursive=recursive))


def find_finished_tournament_files(folder):
    """Fichiers .tournoi terminés (get_live_status()['finished']) trouvés
    directement dans `folder` (non récursif, comme le Lobby SNG) — liste
    de dicts {path, name}, utilisée pour "Archiver les terminés..."."""
    results = []
    for path in find_tournament_files(folder, recursive=False):
        try:
            db = Database(path, read_only=True)
            status = db.get_live_status()
            db.close()
        except Exception:
            continue
        if status["finished"]:
            results.append({"path": path, "name": status["name"]})
    return results


def archive_tournament_files(paths):
    """Déplace chaque fichier .tournoi de `paths` dans un sous-dossier
    "archive" créé (si besoin) dans son propre dossier parent — celui où
    il a été créé, pas un emplacement d'archive centralisé. En cas de nom
    déjà présent dans ce sous-dossier, ajoute un suffixe numérique plutôt
    que d'écraser. Renvoie le nombre de fichiers effectivement déplacés ;
    une erreur sur un fichier (verrouillé, permissions...) n'interrompt
    pas le traitement des autres."""
    moved = 0
    for path in paths:
        folder = os.path.dirname(path)
        archive_dir = os.path.join(folder, "archive")
        try:
            os.makedirs(archive_dir, exist_ok=True)
            base, ext = os.path.splitext(os.path.basename(path))
            dest = os.path.join(archive_dir, base + ext)
            i = 2
            while os.path.exists(dest):
                dest = os.path.join(archive_dir, f"{base}_{i}{ext}")
                i += 1
            shutil.move(path, dest)
            moved += 1
        except OSError:
            continue
    return moved


def _read_tournament_date_ro(path):
    """Comme Database.get_tournament_date, mais en lecture seule (URI
    mode=ro) pour ne jamais créer ni modifier le fichier consulté — même
    précaution que read_player_names_from_file, appliquée ici pour pouvoir
    comparer les dates de plusieurs tournois (dossier entier) sans toucher
    aux fichiers des autres soirées."""
    try:
        uri = f"file:{os.path.abspath(path)}?mode=ro"
        conn = sqlite3.connect(uri, uri=True)
    except sqlite3.OperationalError:
        return ""
    try:
        row = conn.execute(
            "SELECT value FROM settings WHERE key='tournament_date'"
        ).fetchone()
        if row and row[0]:
            return row[0]
        st = os.stat(path)
        ts = getattr(st, "st_birthtime", None)
        if ts is None:
            ts = st.st_mtime
        return time.strftime("%Y-%m-%d", time.localtime(ts))
    except sqlite3.OperationalError:
        return ""
    finally:
        conn.close()


def _player_active_in_file(path, name_lower):
    """Vrai si un joueur nommé `name_lower` (déjà en minuscules) est
    actuellement 'active' dans le fichier .tournoi `path`, consulté en
    lecture seule (même précaution que _read_tournament_date_ro)."""
    try:
        uri = f"file:{os.path.abspath(path)}?mode=ro"
        conn = sqlite3.connect(uri, uri=True)
    except sqlite3.OperationalError:
        return False
    try:
        row = conn.execute(
            "SELECT 1 FROM players WHERE status='active' AND lower(name)=? LIMIT 1",
            (name_lower,),
        ).fetchone()
        return row is not None
    except sqlite3.OperationalError:
        return False
    finally:
        conn.close()


def _active_player_names_lower_in_file(path):
    """Ensemble des noms (en minuscules) actuellement 'active' dans le
    fichier .tournoi `path`, consulté en lecture seule."""
    try:
        uri = f"file:{os.path.abspath(path)}?mode=ro"
        conn = sqlite3.connect(uri, uri=True)
    except sqlite3.OperationalError:
        return set()
    try:
        rows = conn.execute("SELECT lower(name) FROM players WHERE status='active'").fetchall()
        return {r[0] for r in rows}
    except sqlite3.OperationalError:
        return set()
    finally:
        conn.close()


def find_players_active_elsewhere(folder, names, exclude_path=None, date=None):
    """Parmi `names`, renvoie l'ensemble de ceux actuellement actifs dans
    un fichier .tournoi du dossier `folder` (recherche non récursive, hors
    `exclude_path`) — utilisé pour griser, dans la fenêtre "Joueurs
    participants", les joueurs déjà engagés dans un autre tournoi en cours
    du même dossier (ex : un autre Sit & Go). Ne modifie aucun fichier.

    Si `date` (AAAA-MM-JJ) est fourni, seuls les tournois datés du même
    jour sont comparés : un ancien fichier abandonné (jamais terminé, un
    autre jour) ne doit pas griser un joueur indéfiniment."""
    if not folder or not names:
        return set()
    files = [
        p for p in find_tournament_files(folder, recursive=False)
        if not exclude_path or os.path.abspath(p) != os.path.abspath(exclude_path)
    ]
    if date:
        files = [p for p in files if _read_tournament_date_ro(p) == date]
    if not files:
        return set()
    wanted_by_lower = {n.strip().lower(): n for n in names}
    conflicted = set()
    for path in files:
        for name_lower in _active_player_names_lower_in_file(path):
            original = wanted_by_lower.get(name_lower)
            if original:
                conflicted.add(original)
    return conflicted


def _active_players_and_name_in_file(path):
    """(nom_du_tournoi, [noms de joueurs actifs]) pour le fichier .tournoi
    `path`, consulté en lecture seule."""
    try:
        uri = f"file:{os.path.abspath(path)}?mode=ro"
        conn = sqlite3.connect(uri, uri=True)
    except sqlite3.OperationalError:
        return ("", [])
    try:
        name_row = conn.execute(
            "SELECT value FROM settings WHERE key='tournament_name'"
        ).fetchone()
        tournament_name = (
            name_row[0] if name_row and name_row[0]
            else os.path.splitext(os.path.basename(path))[0]
        )
        rows = conn.execute("SELECT name FROM players WHERE status='active'").fetchall()
        return (tournament_name, [r[0] for r in rows])
    except sqlite3.OperationalError:
        return ("", [])
    finally:
        conn.close()


def find_stale_active_players(folder, before_date=None, recursive=True):
    """Parcourt les fichiers .tournoi de `folder` (et ses sous-dossiers si
    `recursive`), et renvoie ceux qui ont encore des joueurs 'active' —
    utilisé par le bouton "Tout réactiver" du répertoire pour repérer les
    joueurs restés coincés dans de vieux tournois abandonnés (jamais
    terminés). Si `before_date` (AAAA-MM-JJ) est fourni, seuls les
    tournois datés strictement avant sont pris en compte ; par défaut
    (None), TOUS les tournois du dossier sont vérifiés, quelle que soit
    leur date — la fenêtre de confirmation affichée avant application
    (voir RosterManagerDialog._reactivate_all) reste le seul garde-fou.
    Lecture seule (aucun fichier modifié ici). Renvoie une liste de dicts
    {path, tournament_name, players: [noms]}."""
    if not folder:
        return []
    results = []
    for path in find_tournament_files(folder, recursive=recursive):
        if before_date:
            date = _read_tournament_date_ro(path)
            if not date or date >= before_date:
                continue
        tournament_name, names = _active_players_and_name_in_file(path)
        if names:
            results.append({"path": path, "tournament_name": tournament_name, "players": names})
    return results


def withdraw_stale_active_players(stale_entries):
    """Applique la correction repérée par find_stale_active_players : pour
    chaque entrée, ouvre le fichier .tournoi en écriture et retire
    (forfait, via Database.withdraw_player) chacun de ses joueurs encore
    actifs. Renvoie le nombre total de joueurs libérés."""
    total = 0
    for entry in stale_entries:
        db = Database(entry["path"])
        try:
            for p in db.list_players(status="active"):
                db.withdraw_player(p["id"])
                total += 1
        finally:
            db.close()
    return total


def find_previous_tournament_files(current_path, current_date=None, count=1):
    """Renvoie jusqu'à `count` chemins de fichiers .tournoi datés avant
    `current_date`, parmi les fichiers .tournoi présents dans le même
    dossier que `current_path` (recherche non récursive : uniquement ce
    dossier, pas ses sous-dossiers), du plus récent au plus ancien. Les
    autres fichiers ne sont ouverts qu'en lecture seule (voir
    `_read_tournament_date_ro`). Renvoie une liste vide si `current_path`
    n'est pas encore sauvegardé sur disque, si `count` <= 0, ou si aucun
    tournoi antérieur n'est trouvé (dates égales ou manquantes ignorées)."""
    if not current_path or not os.path.exists(current_path) or count <= 0:
        return []
    folder = os.path.dirname(os.path.abspath(current_path)) or "."
    if current_date is None:
        current_date = _read_tournament_date_ro(current_path)
    if not current_date:
        return []

    candidates = []  # (date, path)
    for path in find_tournament_files(folder, recursive=False):
        if os.path.abspath(path) == os.path.abspath(current_path):
            continue
        d = _read_tournament_date_ro(path)
        if not d or d >= current_date:
            continue
        candidates.append((d, path))
    candidates.sort(key=lambda c: c[0], reverse=True)  # plus récent en premier
    return [path for _, path in candidates[:count]]


def find_previous_tournament_file(current_path, current_date=None):
    """Comme `find_previous_tournament_files` mais renvoie uniquement le
    plus récent (ou None) — pratique quand on ne veut vérifier qu'un seul
    tournoi précédent."""
    found = find_previous_tournament_files(current_path, current_date, count=1)
    return found[0] if found else None


def build_period_summary(folder, date_from=None, date_to=None, recursive=True):
    """Parcourt tous les fichiers .tournoi d'un dossier et construit une
    synthèse des résultats pour la période indiquée. `date_from` /
    `date_to` sont des chaînes 'AAAA-MM-JJ' (bornes incluses), ou None
    pour ne pas borner. Renvoie un dict :

      {
        "tournaments": [ {name, date, path, entries, prize_pool, status,
                           winner, bounty_distributed}, ... ],  # triés par date
        "players": [ {name, tournaments_played, wins, best_place,
                       total_cost, total_gain, total_bounty_won,
                       total_points}, ... ],  # triés par total_points décroissant
      }

    "total_bounty_won" (par joueur) et "bounty_distributed" (par tournoi)
    viennent de la MÊME source que l'onglet Primes de chaque tournoi (voir
    Database.get_primes_summary, colonne "Mon Bounty") : nombre
    d'éliminations (kills) × valeur d'un bounty (réglage manuel
    `bounty_amount`, sinon 10×√N points) — PAS l'ancien champ
    `bounty_won`/`bounty` (mécanisme cash/PKO indépendant, plus utilisé par
    l'onglet Primes), qui reste à 0 dès que `bounty_amount` vaut 0 (valeur
    par défaut de ce club, qui ne joue qu'en points).

    "total_points" par joueur = somme, sur toute la période, du TOTAL de
    l'onglet Primes de chaque tournoi joué (Présence + Assiduité +
    Classement + Bounty, en points) — pas un calcul en euros ("total_cost"
    / "total_gain" restent disponibles pour qui en aurait besoin, mais
    n'entrent plus dans "total_points")."""
    tournaments = []
    players = {}

    for path in find_tournament_files(folder, recursive=recursive):
        try:
            db = Database(path, read_only=True)
        except Exception:
            continue
        try:
            date = db.get_tournament_date()
            if date_from and date < date_from:
                continue
            if date_to and date > date_to:
                continue

            name = db.get_setting(
                "tournament_name", os.path.splitext(os.path.basename(path))[0]
            )
            all_players = db.list_players()
            active = [p for p in all_players if p["status"] == "active"]
            finished = len(active) == 1
            winner = active[0]["name"] if finished else "-"
            # Même source que l'onglet Primes de ce tournoi (voir
            # get_primes_summary) : "bo_montant" = bounty en points, "total"
            # = Présence + Assiduité + Classement + Bounty pour ce tournoi.
            primes_by_name = {r["name"]: r for r in db.get_primes_summary()}
            bounty_distributed = sum(r["bo_montant"] for r in primes_by_name.values())
            entries = sum(p["buyin_count"] for p in all_players)
            stats = db.get_stats()
            payouts_by_place = {r["place"]: r["amount"] for r in db.get_payouts_amounts()}

            tournaments.append({
                "name": name,
                "date": date,
                "path": path,
                "entries": entries,
                "prize_pool": stats["prize_pool"],
                "status": "Terminé" if finished else "En cours",
                "winner": winner,
                "bounty_distributed": bounty_distributed,
            })

            buyin_amount = db.get_setting_float("buyin_amount", 0)
            rebuy_amount = db.get_setting_float("rebuy_amount", 0)
            addon_amount = db.get_setting_float("addon_amount", 0)

            for p in all_players:
                place = None
                gain = 0.0
                if p["status"] == "eliminated":
                    place = p["place"]
                    gain = payouts_by_place.get(place, 0.0)
                elif p["status"] == "active" and finished:
                    place = 1
                    gain = payouts_by_place.get(1, 0.0)

                cost = (
                    p["buyin_count"] * buyin_amount
                    + p["rebuy_count"] * rebuy_amount
                    + p["addon_count"] * addon_amount
                )
                prime_row = primes_by_name.get(p["name"])
                bounty_won = prime_row["bo_montant"] if prime_row else 0
                points = prime_row["total"] if prime_row else 0

                agg = players.setdefault(p["name"], {
                    "name": p["name"],
                    "tournaments_played": 0,
                    "wins": 0,
                    "best_place": None,
                    "total_cost": 0.0,
                    "total_gain": 0.0,
                    "total_bounty_won": 0,
                    "total_points": 0,
                })
                agg["tournaments_played"] += 1
                agg["total_cost"] += cost
                agg["total_gain"] += gain
                agg["total_bounty_won"] += bounty_won
                agg["total_points"] += points
                if place == 1:
                    agg["wins"] += 1
                if place is not None and (agg["best_place"] is None or place < agg["best_place"]):
                    agg["best_place"] = place
        finally:
            db.close()

    tournaments.sort(key=lambda t: t["date"])
    players_list = sorted(players.values(), key=lambda a: a["total_points"], reverse=True)
    return {"tournaments": tournaments, "players": players_list}


def _period_range_label(date_from, date_to):
    """Texte lisible de la période couverte par une synthèse
    (build_period_summary), à afficher dans les exports (CSV/Excel/PDF) —
    sans lui, un fichier exporté ne permet pas de savoir, une fois hors de
    l'application, sur quelle période (dates du/au choisies dans l'onglet
    Statistiques) portait son contenu."""
    if not date_from and not date_to:
        return "Période : toutes dates confondues"
    if date_from and date_to:
        return f"Période : du {format_date_fr(date_from)} au {format_date_fr(date_to)}"
    if date_from:
        return f"Période : à partir du {format_date_fr(date_from)}"
    return f"Période : jusqu'au {format_date_fr(date_to)}"


# Colonnes disponibles pour l'export de la synthèse par période, sous la
# forme (clé, en-tête, fonction d'extraction de la valeur à partir d'une
# ligne de tournoi/joueur). Définies une seule fois ici et réutilisées à
# la fois par l'export CSV, l'export Excel et la boîte de dialogue de
# sélection des colonnes (main.py) — ainsi les trois restent toujours en
# phase.
PERIOD_TOURNAMENT_COLUMNS = [
    ("date", "Date", lambda t: format_date_fr(t["date"])),
    ("name", "Tournoi", lambda t: t["name"]),
    ("status", "Statut", lambda t: t["status"]),
    ("entries", "Entrées", lambda t: t["entries"]),
    ("prize_pool", "Prize pool (€)", lambda t: round(t["prize_pool"], 2)),
    ("winner", "Vainqueur", lambda t: t["winner"]),
    ("bounty_distributed", "Primes distribuées (€)", lambda t: t["bounty_distributed"]),
]

PERIOD_PLAYER_COLUMNS = [
    ("name", "Joueur", lambda a: a["name"]),
    ("tournaments_played", "Tournois joués", lambda a: a["tournaments_played"]),
    ("wins", "Victoires", lambda a: a["wins"]),
    ("best_place", "Meilleur Rang", lambda a: a["best_place"]),
    ("total_cost", "Total investi (€)", lambda a: round(a["total_cost"], 2)),
    ("total_gain", "Gains classement (€)", lambda a: round(a["total_gain"], 2)),
    ("total_bounty_won", "Bounty", lambda a: a["total_bounty_won"]),
    ("total_points", "TOTAL Pts", lambda a: a["total_points"]),
]

# Colonnes disponibles pour l'export du classement final nominatif d'UN
# tournoi (menu Fichier > Exporter les résultats...). Le rang du vainqueur
# est un entier (1), pas une chaîne, pour éviter le souci d'alignement
# Excel entre texte et nombres dans la même colonne.
RESULT_COLUMNS = [
    ("rang", "Rang", lambda r: r["rang"]),
    ("name", "Nom", lambda r: r["name"]),
    ("status", "Statut", lambda r: r["status"]),
    ("gain", "Gain (€)", lambda r: round(r["gain"], 2) if r["gain"] else None),
    ("buyin", "Buy-ins", lambda r: r["buyin"]),
    ("rebuy", "Rebuys", lambda r: r["rebuy"]),
    ("addon", "Add-ons", lambda r: r["addon"]),
    ("bounty_won", "Prime gagnée (€)", lambda r: r["bounty_won"]),
]

# Colonnes disponibles pour l'export de la grille de gains telle
# qu'affichée dans l'onglet Gains (place -> pourcentage -> montant, sans
# nom de joueur) — distinct du classement nominatif ci-dessus.
PAYOUT_COLUMNS = [
    ("place", "Place", lambda r: r["place"]),
    ("percentage", "Pourcentage (%)", lambda r: round(r["percentage"], 1)),
    ("amount", "Montant (€)", lambda r: round(r["amount"], 2)),
]

# Colonnes disponibles pour l'export de l'onglet Joueurs, dans le même
# ordre que son tableau (nom, table, siège, chips, achats, prime en jeu,
# statut, rang) — distinct du classement final nominatif (RESULT_COLUMNS,
# qui a le gain et la prime déjà empochée plutôt que la prime en jeu).
PLAYERS_TAB_COLUMNS = [
    ("name", "Nom", lambda p: p["name"]),
    ("table", "Table", lambda p: p["table"]),
    ("seat", "Siège", lambda p: p["seat"]),
    ("chips", "Chips", lambda p: p["chips"]),
    ("buyin", "Buy-in", lambda p: p["buyin"]),
    ("rebuy", "Rebuys", lambda p: p["rebuy"]),
    ("addon", "Add-ons", lambda p: p["addon"]),
    ("bounty", "Prime", lambda p: p["bounty"]),
    ("status", "Statut", lambda p: p["status"]),
    ("rang", "Rang", lambda p: p["rang"]),
    ("elim_time", "Éliminé le", lambda p: format_datetime_fr(p["elim_time"])),
    ("elim_round", "Round", lambda p: p["elim_round"]),
    ("eliminated_by", "Éliminé par", lambda p: p["eliminated_by"]),
]

# Colonnes disponibles pour l'export de l'onglet Primes, dans le même ordre
# que son tableau (nom, présence, assiduité, rang, classement, bounty
# nombre/valeur/montant, total). Voir Database.get_primes_summary.
PRIMES_COLUMNS = [
    ("name", "Joueur", lambda r: r["name"]),
    ("rang", "Rang", lambda r: r["rang"]),
    ("presence", "Présence", lambda r: r["presence"]),
    ("assiduite", "Assiduité", lambda r: r["assiduite"]),
    ("cl_montant", "Classement", lambda r: r["cl_montant"]),
    ("bo_nombre", "Nb Bounty", lambda r: r["bo_nombre"]),
    ("bo_valeur", "Val Bounty", lambda r: r["bo_valeur"]),
    ("bo_montant", "Mon Bounty", lambda r: r["bo_montant"]),
    ("total", "TOTAL", lambda r: r["total"]),
]

# Réglages imprimables de l'onglet Paramètres (bouton "Imprimer
# Paramètres...", voir App._print_settings_pdf) : (clé, libellé), dans le
# même ordre que le formulaire (colonne gauche puis colonne droite). Tenu
# à jour manuellement en phase avec main.py._build_settings_tab — pas de
# source unique automatique ici, ces libellés n'existent que côté widgets
# Tk (self.settings_vars), pas dans ce module.
SETTINGS_PRINT_FIELDS = [
    ("club_name", "Nom du Club"),
    ("tournament_name", "Nom du tournoi"),
    ("buyin_amount", "Montant du buy-in (€)"),
    ("rebuy_amount", "Montant d'un rebuy (€)"),
    ("addon_amount", "Montant d'un add-on (€)"),
    ("starting_chips", "Tapis de départ (chips)"),
    ("rebuy_chips", "Chips reçues pour un rebuy"),
    ("addon_chips", "Chips reçues pour un add-on"),
    ("max_seats_per_table", "Nombre de sièges par table"),
    ("min_players_per_table", "Nombre minimum de joueurs par table avant rééquilibrage"),
    ("highlight_duration_minutes", "Durée de surbrillance des derniers joueurs déplacés (minutes)"),
    ("rake_percent", "Rake / frais d'organisation (%)"),
    ("movement_signal_duration_ms", "Durée max. du signal de mouvement (ms)"),
    ("tournament_day_folder", "Chemin du dossier du tournoi du jour"),
    ("start_small_blind", "Small blind (niveau 1)"),
    ("start_big_blind", "Big blind (niveau 1)"),
    ("ante_start_level", "Niveau à partir duquel l'ante commence"),
    ("start_ante", "Valeur de l'ante de départ"),
    ("round_duration_minutes", "Durée d'un Round (minutes)"),
    ("break_duration_minutes", "Durée de la Pause (minutes)"),
    ("attendance_bonus_points", "Prime de présence (points)"),
    ("assiduity_bonus_points", "Prime d'assiduité (points)"),
    ("assiduity_consecutive_days", "Nombre de jours consécutifs (assiduité)"),
    ("ranking_bonus_points", "Prime de classement (points)"),
    ("bounty_amount", "Montant du bounty (points)"),
    ("pko_mode", "Mode PKO (prime progressive)"),
    ("pko_cash_percent", "Part en Perso immédiat en PKO (%)"),
]

# Colonnes disponibles pour l'export de la 2e table de l'onglet Primes,
# l'historique du bounty progressif (mécanisme PKO interne, voir
# get_bounty_events) — distinct du récapitulatif ci-dessus.
BOUNTY_HISTORY_COLUMNS = [
    ("time", "Heure", lambda r: format_datetime_fr(r["event_time"])),
    ("eliminated", "Joueur éliminé", lambda r: r["eliminated_name"]),
    ("eliminator", "Éliminé par", lambda r: r["eliminator_name"] or "-"),
    ("amount", "Points gagnés", lambda r: r["amount_won"]),
    ("grow", "Ajouté à sa prime", lambda r: r["added_to_eliminator_bounty"] or 0),
]


def _selected_period_columns(columns, keys):
    """Sous-ensemble de `columns` (une des listes ci-dessus) correspondant
    à `keys`, dans l'ordre d'origine ; toutes les colonnes si `keys` est
    None."""
    if keys is None:
        return columns
    keys = set(keys)
    return [c for c in columns if c[0] in keys]


def export_period_summary_csv(
    summary, path, tournament_keys=None, player_keys=None, date_from=None, date_to=None,
):
    """Exporte une synthèse (issue de build_period_summary) en CSV : une
    section 'Tournois de la période', puis une section 'Classement des
    joueurs' incluant les primes (bounty) empochées. `tournament_keys` /
    `player_keys` permettent de ne garder qu'un sous-ensemble de colonnes
    (voir PERIOD_TOURNAMENT_COLUMNS / PERIOD_PLAYER_COLUMNS) ; None = toutes.
    `date_from`/`date_to` (AAAA-MM-JJ ou None) : bornes de la période
    choisies dans l'onglet Statistiques, affichées en clair sous chaque
    titre de section (voir _period_range_label)."""
    import csv

    t_cols = _selected_period_columns(PERIOD_TOURNAMENT_COLUMNS, tournament_keys)
    p_cols = _selected_period_columns(PERIOD_PLAYER_COLUMNS, player_keys)
    period_label = _period_range_label(date_from, date_to)

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        if t_cols:
            writer.writerow(["Tournois de la période"])
            writer.writerow([period_label])
            writer.writerow([h for _, h, _ in t_cols])
            for t in summary["tournaments"]:
                writer.writerow([fn(t) for _, _, fn in t_cols])
            writer.writerow([])
        if p_cols:
            writer.writerow(["Classement des joueurs sur la période"])
            writer.writerow([period_label])
            writer.writerow([h for _, h, _ in p_cols])
            for a in summary["players"]:
                writer.writerow([fn(a) for _, _, fn in p_cols])
    return path


def export_period_summary_xlsx(
    summary, path, tournament_keys=None, player_keys=None, date_from=None, date_to=None,
):
    """Exporte une synthèse en Excel (.xlsx) : une feuille 'Tournois', une
    feuille 'Joueurs', avec les mêmes options de sélection de colonnes que
    export_period_summary_csv (voir aussi date_from/date_to là-bas).
    Nécessite le paquet 'openpyxl'."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    t_cols = _selected_period_columns(PERIOD_TOURNAMENT_COLUMNS, tournament_keys)
    p_cols = _selected_period_columns(PERIOD_PLAYER_COLUMNS, player_keys)
    period_label = _period_range_label(date_from, date_to)
    header_fill = PatternFill(start_color="1F4E24", end_color="1F4E24", fill_type="solid")

    def _write_sheet(ws, cols, rows):
        # Ligne 1 : période couverte (italique, fusionnée sur toute la
        # largeur du tableau) ; ligne 2 : en-têtes de colonnes (décalés
        # d'une ligne par rapport à avant, d'où row=2 ci-dessous).
        ws.append([period_label])
        ws.cell(row=1, column=1).font = Font(italic=True, color="555555")
        if len(cols) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.append([h for _, h, _ in cols])
        for col in range(1, len(cols) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([fn(row) for _, _, fn in cols])
        for i, _ in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 20

    wb = Workbook()
    first = True
    if t_cols:
        ws_t = wb.active
        ws_t.title = "Tournois"
        first = False
        _write_sheet(ws_t, t_cols, summary["tournaments"])
    if p_cols:
        ws_p = wb.active if first else wb.create_sheet("Joueurs")
        ws_p.title = "Joueurs"
        _write_sheet(ws_p, p_cols, summary["players"])
    if not t_cols and not p_cols:
        wb.active.title = "Synthèse"

    wb.save(path)
    return path


def export_period_summary_pdf(
    summary, path, tournament_keys=None, player_keys=None, date_from=None, date_to=None,
):
    """Exporte une synthèse en PDF : une page 'Tournois de la période', une
    page 'Classement des joueurs sur la période' (chacune omise si sa
    sélection de colonnes est vide), avec les mêmes options que
    export_period_summary_csv (voir aussi date_from/date_to là-bas).
    Nécessite le paquet 'fpdf2'."""
    from fpdf import FPDF

    t_cols = _selected_period_columns(PERIOD_TOURNAMENT_COLUMNS, tournament_keys)
    p_cols = _selected_period_columns(PERIOD_PLAYER_COLUMNS, player_keys)
    period_label = _period_range_label(date_from, date_to)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    def _draw_section(title, cols, rows):
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, _pdf_text(title), border=0)
        pdf.ln(9)
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(85, 85, 85)
        pdf.cell(0, 7, _pdf_text(period_label), border=0)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(10)

        headers = [h for _, h, _ in cols]
        avail_width = pdf.w - pdf.l_margin - pdf.r_margin
        col_width = avail_width / max(1, len(headers))
        row_height = 7

        header_texts = [_pdf_text(h) for h in headers]
        _pdf_fit_font_size(pdf, header_texts, col_width, bold=True)
        pdf.set_fill_color(31, 78, 36)
        pdf.set_text_color(255, 255, 255)
        for h in header_texts:
            pdf.cell(col_width, row_height + 1, h, border=1, align="C", fill=True)
        pdf.ln(row_height + 1)

        body_texts = [_pdf_text(fn(row)) for row in rows for _, _, fn in cols] or [""]
        _pdf_fit_font_size(pdf, body_texts, col_width, bold=False)
        pdf.set_text_color(0, 0, 0)
        fill_toggle = False
        for row in rows:
            if fill_toggle:
                pdf.set_fill_color(247, 241, 227)
            else:
                pdf.set_fill_color(255, 255, 255)
            for _, _, fn in cols:
                pdf.cell(col_width, row_height, _pdf_text(fn(row)), border=1, align="C", fill=True)
            pdf.ln(row_height)
            fill_toggle = not fill_toggle

    if t_cols:
        _draw_section("Tournois de la période", t_cols, summary["tournaments"])
    if p_cols:
        _draw_section("Classement des joueurs sur la période", p_cols, summary["players"])
    if not t_cols and not p_cols:
        pdf.add_page()  # évite un PDF sans aucune page si tout est décoché

    pdf.output(path)
    return path
